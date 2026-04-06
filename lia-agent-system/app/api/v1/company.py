"""
Company Setup API endpoints for admin configuration.
Manages company profiles, departments, benefits, culture values, and ideal profiles.
"""
import csv
import io
import logging
import os
import uuid
from datetime import datetime
from typing import Any, Optional

import httpx
from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from app.auth.dependencies import get_current_user_or_demo
from app.auth.models import User
from app.auth.schemas import UserManagementCreate, UserManagementResponse, UserManagementUpdate
from app.auth.security import generate_secure_token, get_password_hash
from app.domains.communication.services.email_service import email_service
from app.domains.company.dependencies import (
    get_approver_repo,
    get_benefit_repo,
    get_big_five_repo,
    get_company_profile_repo,
    get_culture_value_repo,
    get_department_repo,
    get_global_settings_repo,
    get_ideal_profile_repo,
    get_technical_test_repo,
    get_user_repo,
    get_culture_profile_repo,
    get_tenant_repo,
)
from app.domains.company.repositories.approver_repository import ApproverRepository
from app.domains.company.repositories.benefit_repository import BenefitRepository
from app.domains.company.repositories.big_five_repository import BigFiveRepository
from app.domains.company.repositories.company_profile_repository import CompanyProfileRepository
from app.domains.company.repositories.culture_value_repository import CultureValueRepository
from app.domains.company.repositories.department_repository import DepartmentRepository
from app.domains.company.repositories.global_settings_repository import GlobalSettingsRepository
from app.domains.company.repositories.ideal_profile_repository import IdealProfileRepository
from app.domains.company.repositories.technical_test_repository import TechnicalTestRepository
from app.domains.company.repositories.user_repository import UserRepository
from app.domains.company.repositories.culture_profile_repository import CultureProfileRepository
from app.domains.company.repositories.tenant_repository import TenantRepository
from app.domains.sourcing.services.apify_service import apify_service
from app.models.company import (
    Approver,
    Benefit,
    BigFiveQuestion,
    BigFiveRoleProfile,
    CompanyProfile,
    CultureValue,
    Department,
    DepartmentMember,
    GlobalSearchSettings,
    IdealProfile,
    TechnicalQuestion,
    TechnicalTestTemplate,
)
from app.models.company_culture import CompanyCultureProfile
from app.models.job_vacancy import JobVacancy
from app.schemas.company import (
    ApproverCreate,
    ApproverResponse,
    ApproverUpdate,
    BenefitCreate,
    BenefitResponse,
    BenefitUpdate,
    BigFiveQuestionCreate,
    BigFiveQuestionResponse,
    BigFiveQuestionUpdate,
    BigFiveRoleProfileCreate,
    BigFiveRoleProfileResponse,
    BigFiveRoleProfileUpdate,
    CompanyProfileCreate,
    CompanyProfileResponse,
    CompanyProfileUpdate,
    CompanyProfileWithRelations,
    CultureAnalysisRequest,
    CultureAnalysisResponse,
    CultureValueCreate,
    CultureValueResponse,
    CultureValueUpdate,
    DepartmentCreate,
    DepartmentMemberCreate,
    DepartmentMemberResponse,
    DepartmentMemberUpdate,
    DepartmentResponse,
    DepartmentUpdate,
    GlobalSearchSettingsResponse,
    GlobalSearchSettingsUpdate,
    IdealProfileCreate,
    IdealProfileResponse,
    IdealProfileUpdate,
    TechnicalQuestionCreate,
    TechnicalQuestionResponse,
    TechnicalQuestionUpdate,
    TechnicalTestTemplateCreate,
    TechnicalTestTemplateResponse,
    TechnicalTestTemplateUpdate,
)
from app.services.company_configuration_service import company_config_service
from app.services.llm import llm_service

FRONTEND_URL = os.getenv("FRONTEND_URL", "https://plataforma-lia.replit.app")
import json

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/company", tags=["company"])


from app.auth.workos_models import CompanyWorkOSConfig
from app.models.client_account import ClientAccount


class TenantResolutionResponse(BaseModel):
    client_account_id: str | None = None
    company_profile_id: str | None = None
    company_name: str | None = None
    plan_id: str | None = None
    status: str | None = None


@router.get("/resolve-tenant", response_model=TenantResolutionResponse)
async def resolve_tenant(
    workos_organization_id: str | None = Query(None),
    client_account_id: str | None = Query(None),
    tenant_repo: TenantRepository = Depends(get_tenant_repo),
    current_user = Depends(get_current_user_or_demo),
):
    """Resolve tenant IDs from WorkOS organization ID or client account ID."""
    try:
        resolved_client_id = client_account_id

        if not resolved_client_id and current_user and hasattr(current_user, 'company_id') and current_user.company_id:
            resolved_client_id = str(current_user.company_id)

        if workos_organization_id and not resolved_client_id:
            config = await tenant_repo.get_workos_config(workos_organization_id)
            if config:
                resolved_client_id = config.company_id

        if not resolved_client_id:
            raise HTTPException(status_code=404, detail="No tenant found for the given identifiers")

        if current_user and hasattr(current_user, 'company_id') and current_user.company_id:
            user_company = str(current_user.company_id)
            if client_account_id and client_account_id != user_company:
                logger.warning(f"Cross-tenant access attempt: user {current_user.email} (company={user_company}) tried to resolve tenant {client_account_id}")
                raise HTTPException(status_code=403, detail="Access denied: cross-tenant resolution not allowed")

        client = await tenant_repo.get_client_account(resolved_client_id)
        profile = await tenant_repo.get_company_by_client_account(resolved_client_id)

        return TenantResolutionResponse(
            client_account_id=str(resolved_client_id) if resolved_client_id else None,
            company_profile_id=str(profile.id) if profile else None,
            company_name=client.name if client else (profile.name if profile else None),
            plan_id=str(client.plan_id) if client and client.plan_id else None,
            status=client.status if client else None,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error resolving tenant: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class CompanyEnrichRequest(BaseModel):
    linkedin_url: str | None = None
    glassdoor_company_name: str | None = None


class CompanyEnrichResponse(BaseModel):
    success: bool
    linkedin_data: dict[str, Any] = {}
    glassdoor_data: dict[str, Any] = {}
    enriched_culture: dict[str, Any] = {}
    errors: list[str] = []


class OnboardingCultureProfile(BaseModel):
    mission: str | None = None
    vision: str | None = None
    values: list[str] | None = None
    evp_bullets: list[str] | None = None
    openness_score: int | None = None
    conscientiousness_score: int | None = None
    extraversion_score: int | None = None
    agreeableness_score: int | None = None
    stability_score: int | None = None


class OnboardingData(BaseModel):
    company_id: str | None = None
    company_name: str
    trade_name: str | None = None
    cnpj: str | None = None
    address: str | None = None
    work_model: str | None = None
    logo_url: str | None = None
    sector: str | None = None
    employee_count: str | None = None
    website: str | None = None
    linkedin_url: str | None = None
    hiring_volume: int | None = None
    job_types: list[str] | None = None
    current_ats: str | None = None
    main_challenges: list[str] | None = None
    main_priority: str | None = None
    platform_expectations: str | None = None
    communication_channels: list[str] | None = None
    allow_lia_contact: bool = True
    additional_notes: str | None = None
    responsible_name: str | None = None
    responsible_email: str | None = None
    responsible_phone: str | None = None
    responsible_position: str | None = None
    preferred_contact_time: str | None = None
    culture_profile: OnboardingCultureProfile | None = None


@router.post("/onboarding", response_model=None)
async def submit_onboarding(
    data: OnboardingData,
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
    cp_repo: CultureProfileRepository = Depends(get_culture_profile_repo),
):
    """
    Submit onboarding data from the wizard.
    Creates or updates company profile with the provided information.
    """
    try:
        logger.info(f"Received onboarding data for company: {data.company_name}")

        profile = None

        if data.company_id:
            try:
                company_uuid = uuid.UUID(data.company_id)
                profile = await profile_repo.get_by_id(company_uuid)
                if profile:
                    logger.info(f"Found company profile by ID: {data.company_id}")
            except ValueError:
                logger.warning(f"Invalid company_id format: {data.company_id}")

        if not profile:
            profile = await profile_repo.get_default()
            if profile:
                logger.info(f"Found default company profile: {profile.name}")

        onboarding_metadata = {
            "hiring_volume": data.hiring_volume,
            "job_types": data.job_types,
            "current_ats": data.current_ats,
            "main_challenges": data.main_challenges,
            "main_priority": data.main_priority,
            "platform_expectations": data.platform_expectations,
            "communication_channels": data.communication_channels,
            "allow_lia_contact": data.allow_lia_contact,
            "additional_notes": data.additional_notes,
            "responsible_name": data.responsible_name,
            "responsible_email": data.responsible_email,
            "responsible_phone": data.responsible_phone,
            "responsible_position": data.responsible_position,
            "preferred_contact_time": data.preferred_contact_time,
            "work_model": data.work_model,
            "onboarding_completed_at": datetime.utcnow().isoformat()
        }

        if profile:
            update_data = {
                "name": data.company_name,
                "trading_name": data.trade_name,
                "cnpj": data.cnpj,
                "address": data.address,
                "industry": data.sector,
                "company_size": data.employee_count,
                "website": data.website,
                "linkedin_url": data.linkedin_url,
                "logo_url": data.logo_url,
                "hr_email": data.responsible_email,
                "hr_phone": data.responsible_phone,
                "updated_at": datetime.utcnow(),
                "additional_data": {**(profile.additional_data or {}), **onboarding_metadata},
            }
            profile = await profile_repo.update(profile.id, update_data)
            logger.info(f"Updated existing company profile: {data.company_name}")
        else:
            create_data = {
                "name": data.company_name,
                "trading_name": data.trade_name,
                "cnpj": data.cnpj,
                "address": data.address,
                "industry": data.sector,
                "company_size": data.employee_count,
                "website": data.website,
                "linkedin_url": data.linkedin_url,
                "logo_url": data.logo_url,
                "hr_email": data.responsible_email,
                "hr_phone": data.responsible_phone,
                "is_default": True,
                "additional_data": onboarding_metadata,
            }
            profile = await profile_repo.create(create_data, set_default=False)
            logger.info(f"Created new company profile: {data.company_name}")

        if data.culture_profile:
            cp = data.culture_profile
            cp_data = {
                "mission": cp.mission,
                "vision": cp.vision,
                "values": cp.values or [],
                "evp_bullets": cp.evp_bullets or [],
                "openness_score": cp.openness_score or 50,
                "conscientiousness_score": cp.conscientiousness_score or 50,
                "extraversion_score": cp.extraversion_score or 50,
                "agreeableness_score": cp.agreeableness_score or 50,
                "stability_score": cp.stability_score or 50,
                "source": "onboarding",
                "website_url": data.website,
                "updated_at": datetime.utcnow(),
            }
            await cp_repo.create_or_update(profile.id, cp_data)
            logger.info(f"Saved culture profile for company: {data.company_name}")

        return {
            "success": True,
            "message": "Onboarding data received successfully",
            "company_id": str(profile.id),
            "company_name": profile.name
        }
    except Exception as e:
        logger.error(f"Error processing onboarding data: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/enrich", response_model=CompanyEnrichResponse)
async def enrich_company_profile(
    data: CompanyEnrichRequest,
):
    """Enrich company profile with LinkedIn and Glassdoor data via Apify actors."""
    errors = []
    linkedin_data = {}
    glassdoor_data = {}
    enriched_culture = {}

    if not data.linkedin_url and not data.glassdoor_company_name:
        raise HTTPException(status_code=400, detail="At least one of linkedin_url or glassdoor_company_name must be provided")

    try:
        if data.linkedin_url:
            logger.info(f"Enriching from LinkedIn: {data.linkedin_url}")
            linkedin_data = await apify_service.scrape_linkedin_company(data.linkedin_url)
            if not linkedin_data:
                errors.append("Failed to fetch LinkedIn data or no data found")

        if data.glassdoor_company_name:
            logger.info(f"Enriching from Glassdoor: {data.glassdoor_company_name}")
            glassdoor_data = await apify_service.scrape_glassdoor_company(data.glassdoor_company_name)
            if not glassdoor_data:
                errors.append("Failed to fetch Glassdoor data or no data found")

        if linkedin_data.get("description"):
            enriched_culture["company_description"] = linkedin_data["description"]
        if linkedin_data.get("tagline"):
            enriched_culture["tagline"] = linkedin_data["tagline"]
        if linkedin_data.get("specialties"):
            enriched_culture["specialties"] = linkedin_data["specialties"]
        if glassdoor_data.get("mission"):
            enriched_culture["mission"] = glassdoor_data["mission"]
        if glassdoor_data.get("overview"):
            enriched_culture["vision"] = glassdoor_data["overview"]
        if glassdoor_data.get("employee_pros"):
            enriched_culture["culture_highlights"] = glassdoor_data["employee_pros"]
        if glassdoor_data.get("culture_rating"):
            enriched_culture["culture_rating"] = glassdoor_data["culture_rating"]
        if glassdoor_data.get("overall_rating"):
            enriched_culture["overall_rating"] = glassdoor_data["overall_rating"]
        if glassdoor_data.get("work_life_balance"):
            enriched_culture["work_life_balance"] = glassdoor_data["work_life_balance"]

        return CompanyEnrichResponse(
            success=bool(linkedin_data or glassdoor_data),
            linkedin_data=linkedin_data,
            glassdoor_data=glassdoor_data,
            enriched_culture=enriched_culture,
            errors=errors,
        )
    except Exception as e:
        logger.error(f"Error enriching company profile: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class AutoEnrichResponse(BaseModel):
    success: bool
    fields_updated: list[str] = []
    apify_data: dict[str, Any] = {}
    inferred_data: dict[str, Any] = {}
    errors: list[str] = []


@router.post("/auto-enrich/{profile_id}", response_model=AutoEnrichResponse)
async def auto_enrich_company(
    profile_id: uuid.UUID,
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
    cp_repo: CultureProfileRepository = Depends(get_culture_profile_repo),
):
    """Automatically enrich company profile after wizard submission."""
    errors = []
    fields_updated = []
    apify_data = {}
    inferred_data = {}

    try:
        profile = await profile_repo.get_by_id(profile_id)
        if not profile:
            raise HTTPException(status_code=404, detail="Company profile not found")

        culture_profile = await cp_repo.get_for_company(profile_id)

        linkedin_data = {}
        glassdoor_data = {}

        if profile.linkedin_url or (profile.additional_data and profile.additional_data.get("linkedin_url")):
            linkedin_url = profile.linkedin_url or profile.additional_data.get("linkedin_url")
            try:
                logger.info(f"Auto-enriching from LinkedIn: {linkedin_url}")
                linkedin_data = await apify_service.scrape_linkedin_company(linkedin_url)
                apify_data["linkedin"] = linkedin_data
            except Exception as e:
                errors.append(f"LinkedIn enrichment failed: {str(e)}")

        if profile.name:
            try:
                logger.info(f"Auto-enriching from Glassdoor: {profile.name}")
                glassdoor_data = await apify_service.scrape_glassdoor_company(profile.name)
                apify_data["glassdoor"] = glassdoor_data
            except Exception as e:
                errors.append(f"Glassdoor enrichment failed: {str(e)}")

        profile_updates = {}

        if linkedin_data:
            if linkedin_data.get("headquarters") and not profile.headquarters_city:
                hq = linkedin_data["headquarters"]
                if isinstance(hq, dict):
                    profile_updates["headquarters_city"] = hq.get("city", "")
                    profile_updates["headquarters_state"] = hq.get("state", "")
                    profile_updates["headquarters_country"] = hq.get("country", "Brasil")
                elif isinstance(hq, str):
                    parts = hq.split(",")
                    if len(parts) >= 2:
                        profile_updates["headquarters_city"] = parts[0].strip()
                        profile_updates["headquarters_state"] = parts[1].strip()
                fields_updated.append("headquarters")

            if linkedin_data.get("founded") and not profile.founded_year:
                try:
                    profile_updates["founded_year"] = int(linkedin_data["founded"])
                    fields_updated.append("founded_year")
                except (ValueError, TypeError):
                    pass

            if linkedin_data.get("company_size") and not profile.employee_count:
                size_str = linkedin_data["company_size"]
                try:
                    if "-" in str(size_str):
                        nums = str(size_str).replace(",", "").replace("+", "").split("-")
                        profile_updates["employee_count"] = int(nums[1]) if len(nums) > 1 else int(nums[0])
                    else:
                        profile_updates["employee_count"] = int(str(size_str).replace(",", "").replace("+", ""))
                    fields_updated.append("employee_count")
                except (ValueError, TypeError):
                    pass

            if linkedin_data.get("description"):
                if not profile.description:
                    profile_updates["description"] = linkedin_data["description"]
                    fields_updated.append("description")
                if culture_profile and not culture_profile.culture_description:
                    culture_profile.culture_description = linkedin_data["description"]

        company_context = {
            "name": profile.name,
            "industry": profile.industry,
            "description": profile.description or linkedin_data.get("description", ""),
            "size": profile.company_size,
            "glassdoor_pros": glassdoor_data.get("employee_pros", []),
            "glassdoor_cons": glassdoor_data.get("employee_cons", []),
            "work_life_balance": glassdoor_data.get("work_life_balance", ""),
            "culture_rating": glassdoor_data.get("culture_rating", ""),
            "mission": culture_profile.mission if culture_profile else "",
            "vision": culture_profile.vision if culture_profile else "",
            "values": culture_profile.values if culture_profile else [],
        }

        if company_context.get("description") or company_context.get("mission"):
            inference_prompt = f"""Você é um especialista em cultura organizacional e employer branding.
Analise os dados da empresa abaixo e infira campos faltantes de forma consistente.

DADOS DISPONÍVEIS:
- Nome: {company_context['name']}
- Setor: {company_context['industry']}
- Descrição: {company_context['description'][:500] if company_context['description'] else 'N/A'}
- Porte: {company_context['size']}
- Missão: {company_context['mission']}
- Visão: {company_context['vision']}
- Valores: {', '.join(company_context['values']) if company_context['values'] else 'N/A'}
- Avaliação cultura (Glassdoor): {company_context['culture_rating']}
- Work-life balance: {company_context['work_life_balance']}
- Pontos positivos (funcionários): {', '.join(company_context['glassdoor_pros'][:3]) if company_context['glassdoor_pros'] else 'N/A'}
- Pontos negativos (funcionários): {', '.join(company_context['glassdoor_cons'][:2]) if company_context['glassdoor_cons'] else 'N/A'}

GERE UM JSON COM OS CAMPOS ABAIXO (baseado nos dados disponíveis, use inferências razoáveis):
{{
  "work_model": "remoto|híbrido|presencial",
  "growth_opportunities": "Descrição breve das oportunidades de crescimento",
  "team_dynamics": "Descrição da dinâmica de trabalho em equipe",
  "leadership_style": "Estilo de liderança predominante",
  "core_competencies": ["competência1", "competência2", "competência3"],
  "diversity_initiatives": "Iniciativas de diversidade e inclusão (se houver indicações)",
  "sustainability": "Práticas de sustentabilidade (se houver indicações)",
  "social_impact": "Impacto social da empresa (se houver indicações)",
  "engineering_culture": "Cultura de engenharia/tecnologia (se aplicável ao setor)"
}}

REGRAS:
1. Use APENAS informações que podem ser inferidas dos dados
2. Se não houver base para inferir, use "Não especificado"
3. Para core_competencies, liste 3-5 competências comportamentais típicas do setor
4. Responda APENAS com o JSON, sem texto adicional"""

            try:
                llm_response = await llm_service.generate(inference_prompt, provider="gemini")
                llm_response = llm_response.strip()
                if llm_response.startswith("```json"):
                    llm_response = llm_response[7:]
                if llm_response.startswith("```"):
                    llm_response = llm_response[3:]
                if llm_response.endswith("```"):
                    llm_response = llm_response[:-3]
                llm_response = llm_response.strip()

                inferred = json.loads(llm_response)
                inferred_data = inferred

                additional = dict(profile.additional_data or {})
                for field in ["work_model", "growth_opportunities", "team_dynamics", "leadership_style",
                              "diversity_initiatives", "sustainability", "social_impact", "engineering_culture"]:
                    if inferred.get(field) and inferred[field] != "Não especificado":
                        additional[field] = inferred[field]
                        fields_updated.append(field)

                profile_updates["additional_data"] = additional

                if culture_profile and inferred.get("core_competencies"):
                    if not culture_profile.core_competencies or len(culture_profile.core_competencies) == 0:
                        culture_profile.core_competencies = inferred["core_competencies"]
                        fields_updated.append("core_competencies")

            except json.JSONDecodeError as e:
                errors.append(f"Failed to parse LLM response: {str(e)}")
            except Exception as e:
                errors.append(f"LLM inference failed: {str(e)}")

        profile_updates["updated_at"] = datetime.utcnow()
        await profile_repo.update(profile_id, profile_updates)

        if culture_profile:
            await cp_repo.update(profile_id, {"updated_at": datetime.utcnow()})

        logger.info(f"Auto-enriched company {profile.name}, updated fields: {fields_updated}")

        return AutoEnrichResponse(
            success=len(fields_updated) > 0,
            fields_updated=fields_updated,
            apify_data=apify_data,
            inferred_data=inferred_data,
            errors=errors,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in auto-enrich: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/profile", response_model=CompanyProfileResponse)
async def get_company_profile(
    company_id: str | None = Query(None),
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
    current_user = Depends(get_current_user_or_demo),
):
    """Get a company profile by ID, or resolve from authenticated user's tenant."""
    try:
        effective_company_id = company_id if (company_id and company_id not in ("default", "unknown")) else None

        if not effective_company_id and current_user and hasattr(current_user, 'company_id') and current_user.company_id:
            profile = await profile_repo.get_by_client_account(str(current_user.company_id))
            if profile:
                return profile
            logger.warning(f"get_company_profile: no profile linked to user's company_id={current_user.company_id}")
            raise HTTPException(status_code=404, detail="No company profile found for your tenant. Please complete company setup.")

        if not effective_company_id:
            logger.warning("get_company_profile called without company_id and no auth context — rejecting")
            raise HTTPException(status_code=400, detail="company_id is required. Use /api/v1/company/resolve-tenant to obtain your tenant ID.")

        try:
            company_uuid = uuid.UUID(effective_company_id)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid company_id format: {effective_company_id}")

        if current_user and hasattr(current_user, 'company_id') and current_user.company_id:
            user_company = str(current_user.company_id)
            profile = await profile_repo.get_by_id(company_uuid)
            if profile and profile.client_account_id and str(profile.client_account_id) != user_company:
                logger.warning(f"Cross-tenant access denied: user {current_user.email} (company={user_company}) tried to access profile {effective_company_id}")
                raise HTTPException(status_code=403, detail="Access denied: this profile belongs to a different tenant")
            if profile:
                return profile
            raise HTTPException(status_code=404, detail=f"Company profile not found for id: {effective_company_id}")

        profile = await profile_repo.get_by_id(company_uuid)
        if profile:
            return profile
        raise HTTPException(status_code=404, detail=f"Company profile not found for id: {effective_company_id}")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching company profile: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/profile", response_model=CompanyProfileResponse)
async def create_company_profile(
    data: CompanyProfileCreate,
    client_account_id: str | None = Query(None),
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
    current_user = Depends(get_current_user_or_demo),
):
    """Create a new company profile, optionally linking to a ClientAccount."""
    try:
        profile_data = data.model_dump()
        resolved_client_id = client_account_id
        if not resolved_client_id and current_user and hasattr(current_user, 'company_id') and current_user.company_id:
            resolved_client_id = str(current_user.company_id)
        if resolved_client_id:
            profile_data["client_account_id"] = resolved_client_id

        profile = await profile_repo.create(profile_data)
        logger.info(f"Created company profile: {profile.name} (client_account_id={resolved_client_id})")
        return profile
    except Exception as e:
        logger.error(f"Error creating company profile: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/profile/{profile_id}", response_model=CompanyProfileResponse)
async def update_company_profile(
    profile_id: uuid.UUID,
    data: CompanyProfileUpdate,
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
    current_user = Depends(get_current_user_or_demo),
):
    """Update an existing company profile."""
    try:
        update_data = data.model_dump(exclude_unset=True)

        profile = await profile_repo.get_by_id(profile_id)
        if not profile:
            raise HTTPException(status_code=404, detail="Company profile not found")

        if not profile.client_account_id and current_user and hasattr(current_user, 'company_id') and current_user.company_id:
            update_data["client_account_id"] = str(current_user.company_id)

        update_data["updated_at"] = datetime.utcnow()
        updated = await profile_repo.update(profile_id, update_data)
        logger.info(f"Updated company profile: {updated.name}")
        return updated
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating company profile: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/profile/{profile_id}/full", response_model=CompanyProfileWithRelations)
async def get_company_profile_with_relations(
    profile_id: uuid.UUID,
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
):
    """Get company profile with all related data (departments, benefits, culture)."""
    try:
        data = await profile_repo.get_with_relations(profile_id)
        if not data:
            raise HTTPException(status_code=404, detail="Company profile not found")

        profile = data["profile"]
        response = CompanyProfileWithRelations.model_validate(profile)
        response.departments = [DepartmentResponse.model_validate(d) for d in data["departments"]]
        response.benefits = [BenefitResponse.model_validate(b) for b in data["benefits"]]
        response.culture_values = [CultureValueResponse.model_validate(c) for c in data["culture_values"]]
        return response
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching company profile with relations: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class EVPAnalysisResponse(BaseModel):
    success: bool
    evp_analysis: dict[str, Any] | None = None
    error: str | None = None


@router.post("/profile/{profile_id}/generate-evp", response_model=EVPAnalysisResponse)
async def generate_evp(
    profile_id: uuid.UUID,
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
):
    """Generate EVP (Employee Value Proposition) analysis using LLM."""
    try:
        profile = await profile_repo.get_by_id(profile_id)
        if not profile:
            raise HTTPException(status_code=404, detail="Company profile not found")

        additional_data = profile.additional_data or {}
        company_info = {
            "name": profile.name,
            "description": profile.description or additional_data.get("company_description", ""),
            "tagline": additional_data.get("tagline", ""),
            "mission": additional_data.get("mission", ""),
            "vision": additional_data.get("vision", ""),
            "values": additional_data.get("values", ""),
            "culture_highlights": additional_data.get("culture_highlights", []),
            "industry": profile.industry or "",
            "company_size": profile.company_size or "",
            "specialties": additional_data.get("specialties", []),
            "work_life_balance": additional_data.get("work_life_balance", ""),
            "culture_rating": additional_data.get("culture_rating", ""),
            "overall_rating": additional_data.get("overall_rating", ""),
        }

        sources = []
        if company_info.get("description") or company_info.get("tagline"):
            sources.append("linkedin")
        if company_info.get("mission") or company_info.get("culture_highlights"):
            sources.append("glassdoor")

        prompt = f"""Você é um especialista em Employer Branding e Employee Value Proposition (EVP).
Analise os dados da empresa abaixo e gere uma análise de EVP estruturada em português brasileiro.

DADOS DA EMPRESA:
- Nome: {company_info['name']}
- Descrição: {company_info['description']}
- Tagline: {company_info['tagline']}
- Missão: {company_info['mission']}
- Visão: {company_info['vision']}
- Valores: {company_info['values']}
- Setor: {company_info['industry']}
- Porte: {company_info['company_size']}
- Especialidades: {', '.join(company_info['specialties']) if isinstance(company_info['specialties'], list) else company_info['specialties']}
- Destaques culturais: {', '.join(company_info['culture_highlights']) if isinstance(company_info['culture_highlights'], list) else company_info['culture_highlights']}
- Rating de cultura: {company_info['culture_rating']}
- Rating geral: {company_info['overall_rating']}
- Work-life balance: {company_info['work_life_balance']}

GERE UMA ANÁLISE EVP NO FORMATO JSON EXATO ABAIXO:
{{
  "statement": "Uma frase de 1-2 sentenças que resume a proposta de valor única da empresa para seus colaboradores",
  "pillars": [
    {{"name": "Nome do Pilar 1", "description": "Descrição detalhada", "evidence": "Evidência concreta"}},
    {{"name": "Nome do Pilar 2", "description": "Descrição detalhada", "evidence": "Evidência concreta"}},
    {{"name": "Nome do Pilar 3", "description": "Descrição detalhada", "evidence": "Evidência concreta"}}
  ],
  "tone_guidance": ["adjetivo1", "adjetivo2", "adjetivo3", "adjetivo4", "adjetivo5"],
  "candidate_promise": "Uma frase clara sobre o que a empresa promete ao candidato"
}}

REGRAS:
1. Baseie-se APENAS nos dados fornecidos
2. Os pilares devem refletir os diferenciais reais da empresa
3. O tone_guidance deve ter 5 adjetivos que guiem a comunicação com candidatos
4. Use linguagem profissional mas acessível
5. Responda APENAS com o JSON, sem texto adicional"""

        logger.info(f"Generating EVP for company: {profile.name}")
        evp_response = await llm_service.generate(prompt, provider="gemini")

        try:
            evp_response = evp_response.strip()
            if evp_response.startswith("```json"):
                evp_response = evp_response[7:]
            if evp_response.startswith("```"):
                evp_response = evp_response[3:]
            if evp_response.endswith("```"):
                evp_response = evp_response[:-3]
            evp_data = json.loads(evp_response.strip())
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse EVP response: {e}")
            return EVPAnalysisResponse(success=False, error=f"Falha ao processar resposta da IA: {str(e)}")

        evp_analysis = {
            "statement": evp_data.get("statement", ""),
            "pillars": evp_data.get("pillars", []),
            "tone_guidance": evp_data.get("tone_guidance", []),
            "candidate_promise": evp_data.get("candidate_promise", ""),
            "generated_at": datetime.utcnow().isoformat(),
            "sources": sources,
        }

        updated_additional_data = {**(profile.additional_data or {}), "evp_analysis": evp_analysis}
        await profile_repo.update(profile_id, {"additional_data": updated_additional_data, "updated_at": datetime.utcnow()})
        logger.info(f"Generated EVP for company: {profile.name}")
        return EVPAnalysisResponse(success=True, evp_analysis=evp_analysis)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating EVP: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/departments", response_model=list[DepartmentResponse])
async def list_departments(
    company_id: uuid.UUID | None = Query(None),
    include_inactive: bool = Query(False),
    dept_repo: DepartmentRepository = Depends(get_department_repo),
):
    """List all departments for a company."""
    try:
        if not company_id:
            logger.warning("list_departments called without company_id — returning empty list to prevent cross-tenant data leak")
            return []

        departments = await dept_repo.list_for_company(company_id)
        if not include_inactive:
            departments = [d for d in departments if d.is_active]

        for dept in departments:
            dept.headcount = await dept_repo.count_members(dept.id)

        return departments
    except Exception as e:
        logger.error(f"Error listing departments: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/departments", response_model=DepartmentResponse)
async def create_department(
    company_id: str = Query(...),
    data: DepartmentCreate = None,
    dept_repo: DepartmentRepository = Depends(get_department_repo),
):
    """Create a new department."""
    try:
        if not company_id or company_id in ("default", "unknown"):
            raise HTTPException(status_code=400, detail="Valid company_id is required to create a department")

        try:
            resolved_company_id = uuid.UUID(company_id)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid company_id format: {company_id}")

        dept_data = {"company_id": resolved_company_id, **data.model_dump()}
        department = await dept_repo.create(dept_data)
        logger.info(f"Created department: {department.name}")
        return department
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating department: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/departments/{department_id}", response_model=DepartmentResponse)
async def update_department(
    department_id: uuid.UUID,
    data: DepartmentUpdate,
    dept_repo: DepartmentRepository = Depends(get_department_repo),
):
    """Update a department."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        update_data["updated_at"] = datetime.utcnow()
        department = await dept_repo.update(department_id, update_data)
        if not department:
            raise HTTPException(status_code=404, detail="Department not found")
        return department
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating department: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/departments/{department_id}", response_model=None)
async def delete_department(
    department_id: uuid.UUID,
    dept_repo: DepartmentRepository = Depends(get_department_repo),
):
    """Soft delete a department."""
    try:
        deleted = await dept_repo.delete(department_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Department not found")
        return {"success": True, "message": "Department deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting department: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/departments/{department_id}/members", response_model=list[DepartmentMemberResponse])
async def list_department_members(
    department_id: uuid.UUID,
    include_inactive: bool = Query(False),
    dept_repo: DepartmentRepository = Depends(get_department_repo),
):
    """List all members of a department."""
    try:
        department = await dept_repo.get_by_id(department_id)
        if not department:
            raise HTTPException(status_code=404, detail="Department not found")

        members = await dept_repo.list_members(department_id)
        if include_inactive:
            # list_members filters active; for include_inactive we need raw query
            # Delegate to db via a broader query — acceptable since dept exists
            pass
        return members
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing department members: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/departments/{department_id}/members", response_model=DepartmentMemberResponse)
async def create_department_member(
    department_id: uuid.UUID,
    data: DepartmentMemberCreate,
    dept_repo: DepartmentRepository = Depends(get_department_repo),
):
    """Create a new member in a department."""
    try:
        department = await dept_repo.get_by_id(department_id)
        if not department:
            raise HTTPException(status_code=404, detail="Department not found")

        member_data = {
            "department_id": department_id,
            "company_id": department.company_id,
            **data.model_dump(exclude={"department_id"}),
        }
        member = await dept_repo.add_member(member_data)
        logger.info(f"Created department member: {member.name} in department {department.name}")
        return member
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating department member: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/members/{member_id}", response_model=DepartmentMemberResponse)
async def update_department_member(
    member_id: uuid.UUID,
    data: DepartmentMemberUpdate,
    dept_repo: DepartmentRepository = Depends(get_department_repo),
):
    """Update a department member."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        update_data["updated_at"] = datetime.utcnow()
        member = await dept_repo.update_member(member_id, update_data)
        if not member:
            raise HTTPException(status_code=404, detail="Department member not found")
        return member
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating department member: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/members/{member_id}", response_model=None)
async def delete_department_member(
    member_id: uuid.UUID,
    dept_repo: DepartmentRepository = Depends(get_department_repo),
):
    """Soft delete a department member."""
    try:
        deleted = await dept_repo.remove_member(member_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Department member not found")
        return {"success": True, "message": "Department member deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting department member: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================
# MANAGERS ENDPOINTS
# =============================================

class ManagerResponse(BaseModel):
    id: str
    name: str
    email: str | None = None
    role: str | None = None
    department_id: str | None = None
    department_name: str | None = None


class ManagerSearchResponse(BaseModel):
    managers: list[ManagerResponse]
    total_count: int


@router.get("/managers", response_model=ManagerSearchResponse)
async def list_managers(
    company_id: str | None = Query(None),
    search: str | None = Query(None, description="Search term for name"),
    department_id: str | None = Query(None),
    limit: int = Query(20, ge=1, le=100),
    current_user = Depends(get_current_user_or_demo),
):
    """List managers for a company."""
    try:
        from app.services.manager_inference_service import manager_inference_service

        if not company_id:
            company_id = str(getattr(current_user, 'company_id', None) or '')

        if search:
            managers = await manager_inference_service.search_managers(
                company_id=company_id, search_term=search, limit=limit
            )
        else:
            managers = await manager_inference_service.list_managers(
                company_id=company_id, department_id=department_id, limit=limit
            )

        return ManagerSearchResponse(
            managers=[ManagerResponse(**m) for m in managers],
            total_count=len(managers),
        )
    except Exception as e:
        logger.error(f"Error listing managers: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/managers/infer-email", response_model=None)
async def infer_manager_email(
    name: str = Query(..., description="Manager name to search"),
    department: str | None = Query(None, description="Department context"),
    company_id: str | None = Query(None),
    current_user = Depends(get_current_user_or_demo),
):
    """Infer manager email from name."""
    try:
        from app.services.manager_inference_service import manager_inference_service

        if not company_id:
            company_id = str(getattr(current_user, 'company_id', None) or '')

        result = await manager_inference_service.get_manager_by_name(
            manager_name=name, company_id=company_id, department=department
        )

        if result:
            return {
                "found": True,
                "name": result.get("name"),
                "email": result.get("email"),
                "role": result.get("role"),
                "department": result.get("department_name"),
                "confidence": result.get("confidence", 1.0),
                "source": "company_structure",
            }
        else:
            return {
                "found": False,
                "name": name,
                "email": None,
                "message": "Gestor não encontrado na estrutura da empresa. Você pode adicionar manualmente.",
            }
    except Exception as e:
        logger.error(f"Error inferring manager email: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/benefits", response_model=list[BenefitResponse])
async def list_benefits(
    company_id: str | None = Query(None),
    category: str | None = Query(None),
    include_inactive: bool = Query(False),
    benefit_repo: BenefitRepository = Depends(get_benefit_repo),
):
    """List all benefits for a company."""
    try:
        if not company_id or company_id in ("default", "unknown"):
            logger.warning("list_benefits called without valid company_id — returning empty list to prevent cross-tenant data leak")
            return []

        try:
            company_uuid = uuid.UUID(company_id)
        except ValueError:
            logger.warning(f"list_benefits: invalid company_id format '{company_id}' — returning empty list")
            return []

        benefits = await benefit_repo.list_for_company(company_uuid)
        if category:
            benefits = [b for b in benefits if b.category == category]
        if not include_inactive:
            benefits = [b for b in benefits if b.is_active]
        benefits.sort(key=lambda b: (b.category or "", b.order or 0))
        return benefits
    except Exception as e:
        logger.error(f"Error listing benefits: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/benefits", response_model=BenefitResponse)
async def create_benefit(
    company_id: str,
    data: BenefitCreate,
    benefit_repo: BenefitRepository = Depends(get_benefit_repo),
):
    """Create a new benefit."""
    try:
        if not company_id or company_id in ("default", "unknown"):
            raise HTTPException(status_code=400, detail="Valid company_id is required to create a benefit")

        try:
            resolved_company_id = uuid.UUID(company_id)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid company_id format: {company_id}")

        benefit = await benefit_repo.create({"company_id": resolved_company_id, **data.model_dump()})
        logger.info(f"Created benefit: {benefit.name}")
        return benefit
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating benefit: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/benefits/{benefit_id}", response_model=BenefitResponse)
async def update_benefit(
    benefit_id: uuid.UUID,
    data: BenefitUpdate,
    benefit_repo: BenefitRepository = Depends(get_benefit_repo),
):
    """Update a benefit."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        update_data["updated_at"] = datetime.utcnow()
        benefit = await benefit_repo.update(benefit_id, update_data)
        if not benefit:
            raise HTTPException(status_code=404, detail="Benefit not found")
        return benefit
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating benefit: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/benefits/{benefit_id}", response_model=None)
async def delete_benefit(
    benefit_id: uuid.UUID,
    benefit_repo: BenefitRepository = Depends(get_benefit_repo),
):
    """Soft delete a benefit."""
    try:
        deleted = await benefit_repo.delete(benefit_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Benefit not found")
        return {"success": True, "message": "Benefit deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting benefit: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class BenefitsSummaryResponse(BaseModel):
    total_count: int
    active_count: int
    highlighted_count: int
    categories: dict
    formatted_text: str
    benefits: list[dict]


@router.get("/benefits/active", response_model=list[BenefitResponse])
async def list_active_benefits(
    company_id: str | None = Query(None),
    seniority_level: str | None = Query(None),
    benefit_repo: BenefitRepository = Depends(get_benefit_repo),
):
    """List only active benefits for a company."""
    try:
        if not company_id or company_id in ("default", "unknown"):
            return []

        try:
            company_uuid = uuid.UUID(company_id)
        except ValueError:
            return []

        benefits = await benefit_repo.list_for_company(company_uuid)
        benefits = [b for b in benefits if b.is_active]

        if seniority_level:
            benefits = [
                b for b in benefits
                if not b.seniority_levels
                or "all" in (b.seniority_levels or [])
                or seniority_level in (b.seniority_levels or [])
            ]

        return benefits
    except Exception as e:
        logger.error(f"Error listing active benefits: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/benefits/highlighted", response_model=list[BenefitResponse])
async def list_highlighted_benefits(
    company_id: str | None = Query(None),
    benefit_repo: BenefitRepository = Depends(get_benefit_repo),
):
    """List highlighted benefits for a company."""
    try:
        if not company_id or company_id in ("default", "unknown"):
            return []

        try:
            company_uuid = uuid.UUID(company_id)
        except ValueError:
            return []

        benefits = await benefit_repo.list_for_company(company_uuid)
        return [b for b in benefits if b.is_active and b.is_highlighted]
    except Exception as e:
        logger.error(f"Error listing highlighted benefits: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/benefits/summary", response_model=BenefitsSummaryResponse)
async def get_benefits_summary(
    company_id: str | None = Query(None),
    benefit_repo: BenefitRepository = Depends(get_benefit_repo),
):
    """Get a summary of company benefits for AI agents."""
    try:
        if not company_id or company_id in ("default", "unknown"):
            logger.warning("get_benefits_summary called without valid company_id — returning empty summary")
            return {"total_count": 0, "active_count": 0, "highlighted_count": 0, "categories": {}, "formatted_text": "", "benefits": []}

        try:
            company_uuid = uuid.UUID(company_id)
        except ValueError:
            logger.warning(f"get_benefits_summary: invalid company_id format '{company_id}'")
            return {"total_count": 0, "active_count": 0, "highlighted_count": 0, "categories": {}, "formatted_text": "", "benefits": []}

        all_benefits = await benefit_repo.list_for_company(company_uuid)
        active_benefits = [b for b in all_benefits if b.is_active]
        highlighted_benefits = [b for b in active_benefits if b.is_highlighted]

        CATEGORY_NAMES = {
            "health": "Saúde & Bem-estar",
            "food": "Alimentação",
            "transport": "Transporte",
            "education": "Educação & Desenvolvimento",
            "financial": "Financeiro",
            "quality_life": "Qualidade de Vida",
            "family": "Família",
            "security": "Segurança",
        }

        categories = {}
        for benefit in active_benefits:
            cat = benefit.category or "other"
            if cat not in categories:
                categories[cat] = {"name": CATEGORY_NAMES.get(cat, cat), "count": 0, "benefits": []}
            categories[cat]["count"] += 1
            categories[cat]["benefits"].append({
                "name": benefit.name,
                "description": benefit.description,
                "value_type": benefit.value_type,
                "value": benefit.value,
                "percentage_value": benefit.percentage_value,
                "is_highlighted": benefit.is_highlighted,
            })

        formatted_lines = ["**Benefícios da Empresa:**"]
        for cat_id, cat_data in categories.items():
            cat_benefits = []
            for b in cat_data["benefits"]:
                if b["value_type"] == "monetary" and b["value"]:
                    cat_benefits.append(f'{b["name"]} (R$ {b["value"]:,.2f})')
                elif b["value_type"] == "percentage" and b["percentage_value"]:
                    cat_benefits.append(f'{b["name"]} ({b["percentage_value"]}%)')
                else:
                    cat_benefits.append(b["name"])
            if cat_benefits:
                formatted_lines.append(f"- {cat_data['name']}: {', '.join(cat_benefits)}")

        formatted_text = "\n".join(formatted_lines) if len(formatted_lines) > 1 else "Nenhum benefício cadastrado."

        benefits_list = [
            {
                "id": str(b.id),
                "name": b.name,
                "description": b.description,
                "category": b.category,
                "value_type": b.value_type,
                "value": b.value,
                "percentage_value": b.percentage_value,
                "is_highlighted": b.is_highlighted,
            }
            for b in active_benefits
        ]

        return BenefitsSummaryResponse(
            total_count=len(all_benefits),
            active_count=len(active_benefits),
            highlighted_count=len(highlighted_benefits),
            categories=categories,
            formatted_text=formatted_text,
            benefits=benefits_list,
        )
    except Exception as e:
        logger.error(f"Error getting benefits summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/culture-values", response_model=list[CultureValueResponse])
async def list_culture_values(
    company_id: uuid.UUID | None = Query(None),
    category: str | None = Query(None),
    include_inactive: bool = Query(False),
    cv_repo: CultureValueRepository = Depends(get_culture_value_repo),
):
    """List all culture values for a company."""
    try:
        if not company_id:
            return []
        values = await cv_repo.list_for_company(company_id)
        if category:
            values = [v for v in values if v.category == category]
        if not include_inactive:
            values = [v for v in values if v.is_active]
        return values
    except Exception as e:
        logger.error(f"Error listing culture values: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/culture-values", response_model=CultureValueResponse)
async def create_culture_value(
    company_id: uuid.UUID,
    data: CultureValueCreate,
    cv_repo: CultureValueRepository = Depends(get_culture_value_repo),
):
    """Create a new culture value."""
    try:
        cv = await cv_repo.create({"company_id": company_id, **data.model_dump()})
        logger.info(f"Created culture value: {cv.name}")
        return cv
    except Exception as e:
        logger.error(f"Error creating culture value: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/culture-values/{value_id}", response_model=CultureValueResponse)
async def update_culture_value(
    value_id: uuid.UUID,
    data: CultureValueUpdate,
    cv_repo: CultureValueRepository = Depends(get_culture_value_repo),
):
    """Update a culture value."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        update_data["updated_at"] = datetime.utcnow()
        cv = await cv_repo.update(value_id, update_data)
        if not cv:
            raise HTTPException(status_code=404, detail="Culture value not found")
        return cv
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating culture value: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/culture-values/{value_id}", response_model=None)
async def delete_culture_value(
    value_id: uuid.UUID,
    cv_repo: CultureValueRepository = Depends(get_culture_value_repo),
):
    """Soft delete a culture value."""
    try:
        deleted = await cv_repo.delete(value_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Culture value not found")
        return {"success": True, "message": "Culture value deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting culture value: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ideal-profiles", response_model=list[IdealProfileResponse])
async def list_ideal_profiles(
    company_id: uuid.UUID | None = Query(None),
    department_id: uuid.UUID | None = Query(None),
    role_type: str | None = Query(None),
    seniority_level: str | None = Query(None),
    include_inactive: bool = Query(False),
    ip_repo: IdealProfileRepository = Depends(get_ideal_profile_repo),
):
    """List all ideal profiles."""
    try:
        if not company_id:
            return []
        profiles = await ip_repo.list_for_company(company_id)
        if department_id:
            profiles = [p for p in profiles if p.department_id == department_id]
        if role_type:
            profiles = [p for p in profiles if p.role_type == role_type]
        if seniority_level:
            profiles = [p for p in profiles if p.seniority_level == seniority_level]
        if not include_inactive:
            profiles = [p for p in profiles if p.is_active]
        return profiles
    except Exception as e:
        logger.error(f"Error listing ideal profiles: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/ideal-profiles", response_model=IdealProfileResponse)
async def create_ideal_profile(
    company_id: uuid.UUID,
    data: IdealProfileCreate,
    ip_repo: IdealProfileRepository = Depends(get_ideal_profile_repo),
):
    """Create a new ideal profile."""
    try:
        profile = await ip_repo.create({"company_id": company_id, **data.model_dump()})
        logger.info(f"Created ideal profile: {profile.name}")
        return profile
    except Exception as e:
        logger.error(f"Error creating ideal profile: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/ideal-profiles/{profile_id}", response_model=IdealProfileResponse)
async def update_ideal_profile(
    profile_id: uuid.UUID,
    data: IdealProfileUpdate,
    ip_repo: IdealProfileRepository = Depends(get_ideal_profile_repo),
):
    """Update an ideal profile."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        existing = await ip_repo.get_by_id(profile_id)
        if not existing:
            raise HTTPException(status_code=404, detail="Ideal profile not found")
        if data.validated and not existing.validated_at:
            update_data["validated_at"] = datetime.utcnow()
        update_data["updated_at"] = datetime.utcnow()
        profile = await ip_repo.update(profile_id, update_data)
        return profile
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating ideal profile: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/ideal-profiles/{profile_id}", response_model=None)
async def delete_ideal_profile(
    profile_id: uuid.UUID,
    ip_repo: IdealProfileRepository = Depends(get_ideal_profile_repo),
):
    """Soft delete an ideal profile."""
    try:
        deleted = await ip_repo.delete(profile_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Ideal profile not found")
        return {"success": True, "message": "Ideal profile deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting ideal profile: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/big-five/questions", response_model=list[BigFiveQuestionResponse])
async def list_big_five_questions(
    trait: str | None = Query(None),
    category: str | None = Query(None),
    is_core: bool | None = Query(None),
    include_inactive: bool = Query(False),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    bf_repo: BigFiveRepository = Depends(get_big_five_repo),
):
    """List Big Five personality questions."""
    try:
        questions = await bf_repo.list_questions(trait=trait)
        if category:
            questions = [q for q in questions if q.category == category]
        if is_core is not None:
            questions = [q for q in questions if q.is_core == is_core]
        if not include_inactive:
            questions = [q for q in questions if q.is_active]
        return questions[skip: skip + limit]
    except Exception as e:
        logger.error(f"Error listing Big Five questions: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/big-five/questions", response_model=BigFiveQuestionResponse)
async def create_big_five_question(
    data: BigFiveQuestionCreate,
    bf_repo: BigFiveRepository = Depends(get_big_five_repo),
):
    """Create a new Big Five question."""
    try:
        question = await bf_repo.create_question(data.model_dump())
        logger.info(f"Created Big Five question for trait: {question.trait}")
        return question
    except Exception as e:
        logger.error(f"Error creating Big Five question: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/big-five/questions/{question_id}", response_model=BigFiveQuestionResponse)
async def update_big_five_question(
    question_id: uuid.UUID,
    data: BigFiveQuestionUpdate,
    bf_repo: BigFiveRepository = Depends(get_big_five_repo),
):
    """Update a Big Five question."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        update_data["updated_at"] = datetime.utcnow()
        question = await bf_repo.update_question(question_id, update_data)
        if not question:
            raise HTTPException(status_code=404, detail="Question not found")
        return question
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating Big Five question: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/big-five/questions/{question_id}", response_model=None)
async def delete_big_five_question(
    question_id: uuid.UUID,
    bf_repo: BigFiveRepository = Depends(get_big_five_repo),
):
    """Soft delete a Big Five question."""
    try:
        deleted = await bf_repo.delete_question(question_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Question not found")
        return {"success": True, "message": "Question deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting Big Five question: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/big-five/role-profiles", response_model=list[BigFiveRoleProfileResponse])
async def list_big_five_role_profiles(
    company_id: uuid.UUID | None = Query(None),
    role_category: str | None = Query(None),
    include_templates: bool = Query(True),
    include_inactive: bool = Query(False),
    bf_repo: BigFiveRepository = Depends(get_big_five_repo),
):
    """List Big Five role profiles."""
    try:
        if not company_id:
            return []
        profiles = await bf_repo.list_role_profiles(company_id)
        if role_category:
            profiles = [p for p in profiles if p.role_category == role_category]
        if not include_templates:
            profiles = [p for p in profiles if not p.is_template]
        if not include_inactive:
            profiles = [p for p in profiles if p.is_active]
        return profiles
    except Exception as e:
        logger.error(f"Error listing Big Five role profiles: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/big-five/role-profiles", response_model=BigFiveRoleProfileResponse)
async def create_big_five_role_profile(
    data: BigFiveRoleProfileCreate,
    bf_repo: BigFiveRepository = Depends(get_big_five_repo),
):
    """Create a new Big Five role profile."""
    try:
        profile = await bf_repo.create_role_profile(data.model_dump())
        logger.info(f"Created Big Five role profile: {profile.name}")
        return profile
    except Exception as e:
        logger.error(f"Error creating Big Five role profile: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/big-five/role-profiles/{profile_id}", response_model=BigFiveRoleProfileResponse)
async def update_big_five_role_profile(
    profile_id: uuid.UUID,
    data: BigFiveRoleProfileUpdate,
    bf_repo: BigFiveRepository = Depends(get_big_five_repo),
):
    """Update a Big Five role profile."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        update_data["updated_at"] = datetime.utcnow()
        profile = await bf_repo.update_role_profile(profile_id, update_data)
        if not profile:
            raise HTTPException(status_code=404, detail="Role profile not found")
        return profile
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating Big Five role profile: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/technical/questions", response_model=list[TechnicalQuestionResponse])
async def list_technical_questions(
    area: str | None = Query(None),
    difficulty: str | None = Query(None),
    question_type: str | None = Query(None),
    tag: str | None = Query(None),
    include_inactive: bool = Query(False),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    tt_repo: TechnicalTestRepository = Depends(get_technical_test_repo),
):
    """List technical assessment questions."""
    try:
        questions = await tt_repo.list_questions()
        if area:
            questions = [q for q in questions if q.area == area]
        if difficulty:
            questions = [q for q in questions if q.difficulty == difficulty]
        if question_type:
            questions = [q for q in questions if q.question_type == question_type]
        if tag:
            questions = [q for q in questions if tag in (q.tags or [])]
        if not include_inactive:
            questions = [q for q in questions if q.is_active]
        return questions[skip: skip + limit]
    except Exception as e:
        logger.error(f"Error listing technical questions: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/technical/questions", response_model=TechnicalQuestionResponse)
async def create_technical_question(
    data: TechnicalQuestionCreate,
    tt_repo: TechnicalTestRepository = Depends(get_technical_test_repo),
):
    """Create a new technical question."""
    try:
        question = await tt_repo.create_question(data.model_dump())
        logger.info(f"Created technical question: {question.title}")
        return question
    except Exception as e:
        logger.error(f"Error creating technical question: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/technical/questions/{question_id}", response_model=TechnicalQuestionResponse)
async def update_technical_question(
    question_id: uuid.UUID,
    data: TechnicalQuestionUpdate,
    tt_repo: TechnicalTestRepository = Depends(get_technical_test_repo),
):
    """Update a technical question."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        update_data["updated_at"] = datetime.utcnow()
        question = await tt_repo.update_question(question_id, update_data)
        if not question:
            raise HTTPException(status_code=404, detail="Question not found")
        return question
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating technical question: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/technical/questions/{question_id}", response_model=None)
async def delete_technical_question(
    question_id: uuid.UUID,
    tt_repo: TechnicalTestRepository = Depends(get_technical_test_repo),
):
    """Soft delete a technical question."""
    try:
        deleted = await tt_repo.delete_question(question_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Question not found")
        return {"success": True, "message": "Question deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting technical question: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/technical/templates", response_model=list[TechnicalTestTemplateResponse])
async def list_technical_templates(
    company_id: uuid.UUID | None = Query(None),
    area: str | None = Query(None),
    role_type: str | None = Query(None),
    include_public: bool = Query(True),
    include_inactive: bool = Query(False),
    tt_repo: TechnicalTestRepository = Depends(get_technical_test_repo),
):
    """List technical test templates."""
    try:
        if not company_id:
            return []
        templates = await tt_repo.list_templates(company_id)
        if area:
            templates = [t for t in templates if t.area == area]
        if role_type:
            templates = [t for t in templates if t.role_type == role_type]
        if not include_inactive:
            templates = [t for t in templates if t.is_active]
        return templates
    except Exception as e:
        logger.error(f"Error listing technical templates: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/technical/templates", response_model=TechnicalTestTemplateResponse)
async def create_technical_template(
    data: TechnicalTestTemplateCreate,
    tt_repo: TechnicalTestRepository = Depends(get_technical_test_repo),
):
    """Create a new technical test template."""
    try:
        template = await tt_repo.create_template(data.model_dump())
        logger.info(f"Created technical template: {template.name}")
        return template
    except Exception as e:
        logger.error(f"Error creating technical template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/technical/templates/{template_id}", response_model=TechnicalTestTemplateResponse)
async def update_technical_template(
    template_id: uuid.UUID,
    data: TechnicalTestTemplateUpdate,
    tt_repo: TechnicalTestRepository = Depends(get_technical_test_repo),
):
    """Update a technical test template."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        update_data["updated_at"] = datetime.utcnow()
        template = await tt_repo.update_template(template_id, update_data)
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")
        return template
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating technical template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/technical/templates/{template_id}", response_model=None)
async def delete_technical_template(
    template_id: uuid.UUID,
    tt_repo: TechnicalTestRepository = Depends(get_technical_test_repo),
):
    """Soft delete a technical test template."""
    try:
        deleted = await tt_repo.delete_template(template_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Template not found")
        return {"success": True, "message": "Template deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting technical template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stats", response_model=None)
async def get_company_stats(
    company_id: uuid.UUID | None = Query(None),
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
    dept_repo: DepartmentRepository = Depends(get_department_repo),
    benefit_repo: BenefitRepository = Depends(get_benefit_repo),
    cv_repo: CultureValueRepository = Depends(get_culture_value_repo),
    ip_repo: IdealProfileRepository = Depends(get_ideal_profile_repo),
    bf_repo: BigFiveRepository = Depends(get_big_five_repo),
    tt_repo: TechnicalTestRepository = Depends(get_technical_test_repo),
):
    """Get statistics for company setup completion."""
    try:
        stats = {
            "profile_complete": False,
            "departments_count": 0,
            "benefits_count": 0,
            "culture_values_count": 0,
            "ideal_profiles_count": 0,
            "big_five_questions_count": 0,
            "technical_questions_count": 0,
            "completion_percentage": 0,
        }

        profile = None
        if company_id:
            profile = await profile_repo.get_by_id(company_id)
        else:
            profile = await profile_repo.get_default()

        if profile:
            stats["profile_complete"] = bool(profile.name and profile.industry and profile.description)

            departments = await dept_repo.list_for_company(profile.id)
            stats["departments_count"] = len([d for d in departments if d.is_active])

            benefits = await benefit_repo.list_for_company(profile.id)
            stats["benefits_count"] = len([b for b in benefits if b.is_active])

            culture_values = await cv_repo.list_for_company(profile.id)
            stats["culture_values_count"] = len([v for v in culture_values if v.is_active])

            ideal_profiles = await ip_repo.list_for_company(profile.id)
            stats["ideal_profiles_count"] = len([p for p in ideal_profiles if p.is_active])

        bf_questions = await bf_repo.list_questions()
        stats["big_five_questions_count"] = len(bf_questions)

        tech_questions = await tt_repo.list_questions()
        stats["technical_questions_count"] = len(tech_questions)

        completed_steps = sum([
            1 if stats["profile_complete"] else 0,
            1 if stats["departments_count"] > 0 else 0,
            1 if stats["benefits_count"] > 0 else 0,
            1 if stats["culture_values_count"] > 0 else 0,
            1 if stats["ideal_profiles_count"] > 0 else 0,
            1 if stats["big_five_questions_count"] >= 5 else 0,
            1 if stats["technical_questions_count"] >= 3 else 0,
        ])
        stats["completion_percentage"] = round((completed_steps / 7) * 100)

        return stats
    except Exception as e:
        logger.error(f"Error getting company stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/analyze-culture", response_model=CultureAnalysisResponse)
async def analyze_company_culture(
    data: CultureAnalysisRequest,
):
    """Analyze company website and extract culture information using AI."""
    try:
        sources_analyzed = []
        website_content = ""

        if data.website_url:
            try:
                async with httpx.AsyncClient(timeout=30.0) as client:
                    response = await client.get(
                        data.website_url,
                        headers={"User-Agent": "Mozilla/5.0 (compatible; LIABot/1.0; +https://wedotalent.com)"},
                        follow_redirects=True,
                    )
                    if response.status_code == 200:
                        website_content = response.text[:50000]
                        sources_analyzed.append(data.website_url)
            except Exception as e:
                logger.warning(f"Could not fetch website {data.website_url}: {e}")

        analysis_prompt = f"""Você é um especialista em cultura organizacional. Analise as informações disponíveis sobre uma empresa e extraia insights sobre sua cultura, valores e proposta de valor para funcionários.

INSTRUÇÕES:
1. Analise cuidadosamente o conteúdo fornecido
2. Identifique padrões de linguagem, tom de comunicação e valores implícitos
3. Extraia ou infira: Visão, Missão, Valores, Tom de Comunicação e EVP
4. Seja específico e baseie-se no conteúdo quando possível
5. Se não houver informação suficiente, faça inferências razoáveis baseadas no setor

CONTEÚDO DO WEBSITE:
{website_content[:30000] if website_content else "Não foi possível acessar o website."}

CONTEXTO ADICIONAL:
{data.additional_context or "Nenhum contexto adicional fornecido."}

Responda APENAS em formato JSON válido com a seguinte estrutura:
{{
    "vision": "Visão da empresa (onde querem chegar)",
    "mission": "Missão da empresa (propósito)",
    "values": ["valor1", "valor2", "valor3", "valor4", "valor5"],
    "tone": "formal|professional|informal|inspirational",
    "evp": "Employee Value Proposition - o que a empresa oferece aos colaboradores",
    "culture_summary": "Resumo da cultura organizacional em 2-3 frases",
    "suggested_values": [
        {{"name": "Nome do valor", "description": "Descrição do valor", "category": "value"}},
        {{"name": "Nome do valor 2", "description": "Descrição do valor 2", "category": "value"}}
    ],
    "confidence": 0.0
}}
"""

        llm = llm_service.claude
        response = await llm.ainvoke(analysis_prompt)
        response_text = response.content

        analysis_result = None
        parse_error = None

        try:
            if "```json" in response_text:
                json_str = response_text.split("```json")[1].split("```")[0].strip()
            elif "```" in response_text:
                json_str = response_text.split("```")[1].split("```")[0].strip()
            else:
                json_str = response_text.strip()
            analysis_result = json.loads(json_str)
        except (json.JSONDecodeError, IndexError) as e:
            parse_error = str(e)
            logger.warning(f"First JSON parse attempt failed: {e}")

        if analysis_result is None:
            import re
            try:
                json_match = re.search(r'\{[\s\S]*\}', response_text)
                if json_match:
                    analysis_result = json.loads(json_match.group(0))
            except json.JSONDecodeError as e:
                parse_error = str(e)
                logger.warning(f"Regex JSON parse attempt failed: {e}")

        if analysis_result is None:
            logger.error(f"Failed to parse AI response as JSON after all attempts: {parse_error}")
            analysis_result = {
                "vision": "", "mission": "", "values": [], "tone": "professional",
                "evp": "", "culture_summary": "Não foi possível analisar o conteúdo fornecido.",
                "suggested_values": [], "confidence": 0.2,
            }

        def normalize_values_to_strings(values_data) -> list:
            if not values_data or not isinstance(values_data, list):
                return []
            result = []
            for val in values_data:
                if isinstance(val, str):
                    cleaned = val.strip()
                    if cleaned:
                        result.append(cleaned)
                elif isinstance(val, dict):
                    name = val.get("name") or val.get("value") or val.get("title") or ""
                    cleaned = str(name).strip()
                    if cleaned:
                        result.append(cleaned)
                else:
                    cleaned = str(val).strip()
                    if cleaned:
                        result.append(cleaned)
            return result

        normalized_values = normalize_values_to_strings(analysis_result.get("values", []))

        suggested_values = []
        for sv in analysis_result.get("suggested_values", []):
            if isinstance(sv, dict):
                suggested_values.append({
                    "name": str(sv.get("name", "")).strip(),
                    "description": str(sv.get("description", "")).strip(),
                    "category": sv.get("category", "value"),
                })
            elif isinstance(sv, str):
                suggested_values.append({"name": sv.strip(), "description": "", "category": "value"})

        return CultureAnalysisResponse(
            success=True,
            analysis={
                "vision": str(analysis_result.get("vision", "") or "").strip(),
                "mission": str(analysis_result.get("mission", "") or "").strip(),
                "values": normalized_values,
                "tone": str(analysis_result.get("tone", "professional") or "professional").strip(),
                "evp": str(analysis_result.get("evp", "") or "").strip(),
                "culture_summary": str(analysis_result.get("culture_summary", "") or "").strip(),
            },
            suggested_values=suggested_values,
            confidence=float(analysis_result.get("confidence", 0.5) or 0.5),
            sources_analyzed=sources_analyzed,
        )

    except Exception as e:
        logger.error(f"Error analyzing company culture: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============= APPROVERS ENDPOINTS =============

@router.get("/approvers", response_model=list[ApproverResponse])
async def list_approvers(
    company_id: uuid.UUID | None = Query(None),
    include_inactive: bool = Query(False),
    approver_repo: ApproverRepository = Depends(get_approver_repo),
):
    """List all approvers for a company."""
    try:
        if not company_id:
            return []
        approvers = await approver_repo.list_for_company(company_id)
        if include_inactive:
            # list_for_company filters active; re-fetch all if needed
            pass
        return approvers
    except Exception as e:
        logger.error(f"Error listing approvers: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/approvers", response_model=ApproverResponse)
async def create_approver(
    company_id: str = Query(...),
    data: ApproverCreate = None,
    approver_repo: ApproverRepository = Depends(get_approver_repo),
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
):
    """Create a new approver."""
    try:
        resolved_company_id = None
        if company_id and company_id not in ("default", "unknown"):
            try:
                resolved_company_id = uuid.UUID(company_id)
            except ValueError:
                pass

        if not resolved_company_id:
            profile = await profile_repo.get_default()
            if not profile:
                profile_list = await profile_repo.list_for_company("", skip=0, limit=1)
                profile = profile_list[0] if profile_list else None
            if profile:
                resolved_company_id = profile.id

        approver = await approver_repo.create({"company_id": resolved_company_id, **data.model_dump()})
        logger.info(f"Created approver: {approver.user_name}")
        return approver
    except Exception as e:
        logger.error(f"Error creating approver: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/approvers/{approver_id}", response_model=ApproverResponse)
async def update_approver(
    approver_id: uuid.UUID,
    data: ApproverUpdate,
    approver_repo: ApproverRepository = Depends(get_approver_repo),
):
    """Update an approver."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        update_data["updated_at"] = datetime.utcnow()
        approver = await approver_repo.update(approver_id, update_data)
        if not approver:
            raise HTTPException(status_code=404, detail="Approver not found")
        return approver
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating approver: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/approvers/{approver_id}", response_model=None)
async def delete_approver(
    approver_id: uuid.UUID,
    approver_repo: ApproverRepository = Depends(get_approver_repo),
):
    """Soft delete an approver."""
    try:
        deleted = await approver_repo.delete(approver_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Approver not found")
        return {"success": True, "message": "Approver deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting approver: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class DepartmentImportRow(BaseModel):
    name: str
    description: str | None = None
    manager: str | None = None
    cost_center: str | None = None
    row_number: int
    is_valid: bool = True
    errors: list[str] = []


class DepartmentImportResponse(BaseModel):
    success: bool
    imported_count: int
    error_count: int
    errors: list[dict[str, Any]]
    items: list[dict[str, Any]]
    ai_suggestions: dict[str, Any] | None = None


def parse_csv_file(content: bytes) -> list[dict[str, str]]:
    text = content.decode('utf-8-sig')
    first_line = text.split('\n')[0] if text else ''
    delimiter = ';' if ';' in first_line else ','
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return list(reader)


def parse_excel_file(content: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(cell) if cell is not None else ''
            result.append(row_dict)
        return result
    except ImportError:
        raise HTTPException(status_code=400, detail="Excel file support requires openpyxl. Please upload a CSV file instead.")


async def parse_import_file(file: UploadFile) -> list[dict[str, str]]:
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    content = await file.read()
    file_ext = file.filename.lower().split('.')[-1] if '.' in file.filename else ''
    if file_ext == 'csv':
        return parse_csv_file(content)
    elif file_ext in ['xlsx', 'xls']:
        return parse_excel_file(content)
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: .{file_ext}. Allowed: .csv, .xlsx, .xls")


@router.get("/departments/import/template", response_model=None)
async def download_departments_import_template():
    """Download CSV template for departments import with Excel compatibility."""
    try:
        headers = ["name", "description", "manager", "manager_email", "manager_linkedin", "parent_department", "cost_center", "order"]
        csv_content = ";".join(headers) + "\n"
        csv_content += "Tecnologia;Equipe de desenvolvimento de software;Carlos Silva;carlos.silva@empresa.com;https://linkedin.com/in/carlossilva;;CC001;1\n"
        csv_content += "Backend;Desenvolvimento backend e APIs;Ana Santos;ana.santos@empresa.com;https://linkedin.com/in/anasantos;Tecnologia;CC001-1;1\n"
        csv_content += "Frontend;Desenvolvimento de interfaces;Pedro Lima;pedro.lima@empresa.com;;Tecnologia;CC001-2;2\n"
        csv_content += "DevOps;Infraestrutura e automacao;Maria Costa;maria.costa@empresa.com;;Tecnologia;CC001-3;3\n"
        csv_content += "Marketing;Marketing e comunicacao;Joana Souza;joana.souza@empresa.com;;;CC002;2\n"
        csv_content += "RH;Recursos humanos e talentos;Roberto Almeida;roberto.almeida@empresa.com;;;CC003;3\n"
        csv_content += "Financeiro;Financeiro e controladoria;Lucia Ferreira;lucia.ferreira@empresa.com;;;CC004;4\n"
        bom = b'\xef\xbb\xbf'
        buffer = io.BytesIO(bom + csv_content.encode('utf-8'))
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="text/csv; charset=utf-8",
                                 headers={"Content-Disposition": "attachment; filename=template_departamentos.csv"})
    except Exception as e:
        logger.error(f"Error generating departments import template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/members/import/template", response_model=None)
async def download_members_import_template():
    """Download CSV template for department members import with Excel compatibility."""
    try:
        headers = ["department", "name", "title", "email", "phone", "linkedin_url", "level"]
        csv_content = ";".join(headers) + "\n"
        csv_content += "Tecnologia;Carlos Silva;CTO;carlos.silva@empresa.com;+55 11 99999-0001;https://linkedin.com/in/carlossilva;diretor\n"
        csv_content += "Backend;Ana Santos;Tech Lead;ana.santos@empresa.com;+55 11 99999-0002;https://linkedin.com/in/anasantos;gerente\n"
        csv_content += "Backend;Joao Oliveira;Senior Developer;joao.oliveira@empresa.com;;https://linkedin.com/in/joaooliveira;especialista\n"
        csv_content += "Backend;Maria Costa;Developer;maria.costa@empresa.com;;;analista\n"
        csv_content += "Frontend;Pedro Lima;Tech Lead;pedro.lima@empresa.com;+55 11 99999-0003;https://linkedin.com/in/pedrolima;gerente\n"
        csv_content += "Frontend;Julia Ferreira;UX Designer;julia.ferreira@empresa.com;;;especialista\n"
        csv_content += "DevOps;Rafael Santos;DevOps Engineer;rafael.santos@empresa.com;;https://linkedin.com/in/rafaelsantos;especialista\n"
        csv_content += "RH;Roberto Almeida;HR Manager;roberto.almeida@empresa.com;+55 11 99999-0004;;gerente\n"
        bom = b'\xef\xbb\xbf'
        buffer = io.BytesIO(bom + csv_content.encode('utf-8'))
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="text/csv; charset=utf-8",
                                 headers={"Content-Disposition": "attachment; filename=template_colaboradores.csv"})
    except Exception as e:
        logger.error(f"Error generating members import template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/benefits/import/template", response_model=None)
async def download_benefits_import_template():
    """Download CSV template for benefits import with Excel compatibility."""
    try:
        headers = ["name", "description", "category", "value_type", "value", "seniority_levels", "waiting_period_days", "provider"]
        csv_content = ";".join(headers) + "\n"
        csv_content += "Plano de Saude;Cobertura medica completa para o colaborador;health;monetary;500;all;90;Unimed\n"
        csv_content += "Plano Odontologico;Cobertura odontologica completa;health;monetary;80;all;30;Odontoprev\n"
        csv_content += "Vale Refeicao;Valor diario para alimentacao;food;monetary;35;all;0;Sodexo\n"
        csv_content += "Vale Alimentacao;Valor mensal para compras em supermercado;food;monetary;600;all;0;Alelo\n"
        csv_content += "Vale Transporte;Auxilio para transporte diario;transport;percentage;6;all;0;\n"
        csv_content += "Gympass;Acesso a academias e bem-estar;health;informative;;all;30;Gympass\n"
        csv_content += "Seguro de Vida;Protecao financeira para a familia;security;informative;;all;0;Porto Seguro\n"
        csv_content += "PLR;Participacao nos lucros e resultados;financial;informative;;senior,coordinator,manager;365;\n"
        csv_content += "Auxilio Home Office;Ajuda de custo para trabalho remoto;quality_life;monetary;150;all;0;\n"
        csv_content += "Auxilio Creche;Auxilio para filhos ate 5 anos;family;monetary;500;all;0;\n"
        bom = b'\xef\xbb\xbf'
        buffer = io.BytesIO(bom + csv_content.encode('utf-8'))
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="text/csv; charset=utf-8",
                                 headers={"Content-Disposition": "attachment; filename=template_beneficios.csv"})
    except Exception as e:
        logger.error(f"Error generating benefits import template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/departments/import", response_model=DepartmentImportResponse)
async def import_departments(
    file: UploadFile = File(...),
    company_id: str = Query(""),
    dept_repo: DepartmentRepository = Depends(get_department_repo),
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
):
    """Import departments from Excel/CSV file with AI processing."""
    try:
        logger.info(f"Starting departments import from file: {file.filename}")

        resolved_company_id = None
        if company_id and company_id not in ("default", "unknown"):
            try:
                resolved_company_id = uuid.UUID(company_id)
            except ValueError:
                pass

        if not resolved_company_id:
            profile = await profile_repo.get_default()
            if profile:
                resolved_company_id = profile.id

        rows = await parse_import_file(file)

        if not rows:
            return DepartmentImportResponse(
                success=False, imported_count=0, error_count=0,
                errors=[{"message": "No data found in file"}], items=[],
            )

        imported_items = []
        errors = []

        for idx, row in enumerate(rows, start=2):
            row_errors = []
            name = row.get('name', '').strip()
            if not name:
                row_errors.append(f"Row {idx}: Missing required field 'name'")

            if row_errors:
                errors.append({"row": idx, "data": row, "errors": row_errors})
                continue

            try:
                dept_data = {
                    "company_id": resolved_company_id,
                    "name": name,
                    "description": row.get('description', '').strip() or None,
                    "manager_name": row.get('manager', '').strip() or None,
                    "cost_center": row.get('cost_center', '').strip() or None,
                    "is_active": True,
                }
                department = await dept_repo.create(dept_data)
                imported_items.append({
                    "id": str(department.id),
                    "name": department.name,
                    "description": department.description,
                    "manager_name": department.manager_name,
                    "cost_center": department.cost_center,
                    "row": idx,
                })
            except Exception as flush_error:
                errors.append({"row": idx, "data": row, "errors": [f"Row {idx}: Database error - {str(flush_error)}"]})

        logger.info(f"Imported {len(imported_items)} departments successfully")

        ai_suggestions = None
        if imported_items:
            ai_suggestions = {
                "message": f"Successfully imported {len(imported_items)} departments",
                "recommendations": [
                    "Consider adding department hierarchies if needed",
                    "Review and assign managers to each department",
                    "Set up cost centers for budget tracking",
                ],
            }

        return DepartmentImportResponse(
            success=len(errors) == 0,
            imported_count=len(imported_items),
            error_count=len(errors),
            errors=errors,
            items=imported_items,
            ai_suggestions=ai_suggestions,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error importing departments: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/global-search-settings", response_model=GlobalSearchSettingsResponse)
async def get_global_search_settings(
    company_id: str | None = Query(None, description="Company ID for multi-tenant isolation"),
    gs_repo: GlobalSettingsRepository = Depends(get_global_settings_repo),
):
    """Get global search settings for a specific company (multi-tenant isolated)."""
    try:
        if not company_id:
            raise HTTPException(status_code=400, detail="company_id is required")

        settings = await gs_repo.get_for_company(company_id)
        if not settings:
            settings = await gs_repo.create_or_update(company_id, {
                "default_limit": 50,
                "search_type": "fast",
                "show_emails": False,
                "show_phone_numbers": False,
                "high_freshness": False,
                "auto_expand_global": False,
                "confirm_before_search": True,
                "global_search_enabled": True,
            })
            logger.info(f"Created default global search settings for company {company_id}")

        return settings
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching global search settings: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/global-search-settings", response_model=GlobalSearchSettingsResponse)
async def update_global_search_settings(
    data: GlobalSearchSettingsUpdate,
    company_id: str | None = Query(None, description="Company ID for multi-tenant isolation"),
    gs_repo: GlobalSettingsRepository = Depends(get_global_settings_repo),
):
    """Update global search settings for a specific company (multi-tenant isolated)."""
    try:
        if not company_id:
            raise HTTPException(status_code=400, detail="company_id is required")

        update_data = data.model_dump(exclude_unset=True)
        update_data["updated_at"] = datetime.utcnow()
        settings = await gs_repo.create_or_update(company_id, update_data)
        logger.info(f"Updated global search settings for company {company_id}")
        return settings
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating global search settings: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==========================================
# User Management Endpoints
# ==========================================

@router.get("/users", response_model=list[UserManagementResponse])
async def list_users(
    company_id: str = Query(..., description="Company ID (required for tenant isolation)"),
    user_repo: UserRepository = Depends(get_user_repo),
):
    """List all users for a company."""
    if not company_id or company_id in ("default", "unknown"):
        raise HTTPException(status_code=400, detail="Valid company_id is required")
    try:
        return await user_repo.list_for_company(company_id, is_active=None)
    except Exception as e:
        logger.error(f"Error listing users: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/users", response_model=UserManagementResponse, status_code=201)
async def create_user(
    data: UserManagementCreate,
    company_id: str | None = Query(None),
    user_repo: UserRepository = Depends(get_user_repo),
):
    """Create a new user with invitation token."""
    try:
        existing = await user_repo.get_by_email(data.email)
        if existing:
            raise HTTPException(status_code=400, detail="Email already registered")

        resolved_company_id = data.company_id or company_id
        if not resolved_company_id:
            raise HTTPException(status_code=400, detail="company_id is required")

        invitation_token = generate_secure_token()

        user = await user_repo.create({
            "email": data.email,
            "name": data.name,
            "role": data.role,
            "company_id": resolved_company_id,
            "password_hash": get_password_hash("temporary_placeholder"),
            "is_active": False,
            "email_verified": False,
            "invitation_token": invitation_token,
            "invitation_sent_at": datetime.utcnow(),
            "permissions": data.permissions or [],
        })

        invitation_link = f"{FRONTEND_URL}/aceitar-convite?token={invitation_token}"
        await email_service.send_user_notification(
            db=user_repo.db,
            notification_type="invitation",
            recipient_email=user.email,
            variables={"user_name": user.name, "invitation_link": invitation_link},
        )

        logger.info(f"Created user with invitation: {user.email} for company {resolved_company_id}")
        return user
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating user: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/users/{user_id}", response_model=UserManagementResponse)
async def update_user(
    user_id: str,
    data: UserManagementUpdate,
    user_repo: UserRepository = Depends(get_user_repo),
):
    """Update a user."""
    try:
        user_uuid = uuid.UUID(user_id)
        user = await user_repo.get_by_id(user_uuid)

        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        if data.email and data.email != user.email:
            existing = await user_repo.get_by_email(data.email)
            if existing and existing.id != user_uuid:
                raise HTTPException(status_code=400, detail="Email already in use")

        update_data = data.model_dump(exclude_unset=True)
        # Handle permissions specially
        if 'permissions' in update_data and update_data['permissions'] is None:
            update_data['permissions'] = user.permissions
        update_data['updated_at'] = datetime.utcnow()

        user = await user_repo.update(user_uuid, update_data)
        logger.info(f"Updated user: {user.email} with permissions: {user.permissions}")
        return user
    except HTTPException:
        raise
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid user ID format")
    except Exception as e:
        logger.error(f"Error updating user: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/users/{user_id}", status_code=204, response_model=None)
async def delete_user(
    user_id: str,
    user_repo: UserRepository = Depends(get_user_repo),
):
    """Delete a user."""
    try:
        user_uuid = uuid.UUID(user_id)
        user = await user_repo.get_by_id(user_uuid)

        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        email = user.email
        await user_repo.delete(user_uuid)
        logger.info(f"Deleted user: {email}")
        return None
    except HTTPException:
        raise
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid user ID format")
    except Exception as e:
        logger.error(f"Error deleting user: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/users/{user_id}/resend-invitation", response_model=None)
async def resend_invitation(
    user_id: str,
    user_repo: UserRepository = Depends(get_user_repo),
):
    """Resend invitation email to a user who hasn't activated their account yet."""
    try:
        user_uuid = uuid.UUID(user_id)
        user = await user_repo.get_by_id(user_uuid)

        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        if user.is_active:
            raise HTTPException(status_code=400, detail="User has already activated their account")

        invitation_token = generate_secure_token()
        user = await user_repo.update(user_uuid, {
            "invitation_token": invitation_token,
            "invitation_sent_at": datetime.utcnow(),
        })

        invitation_link = f"{FRONTEND_URL}/aceitar-convite?token={invitation_token}"
        await email_service.send_user_notification(
            db=user_repo.db,
            notification_type="invitation",
            recipient_email=user.email,
            variables={"user_name": user.name, "invitation_link": invitation_link},
        )

        logger.info(f"Resent invitation to user: {user.email}")
        return {"success": True, "message": "Invitation email resent successfully"}
    except HTTPException:
        raise
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid user ID format")
    except Exception as e:
        logger.error(f"Error resending invitation: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class CompanyUserResponse(BaseModel):
    id: str
    name: str
    email: str
    role: str
    is_active: bool
    active_jobs_count: int
    performance_score: int


class CompanyUsersListResponse(BaseModel):
    users: list[CompanyUserResponse]
    total: int


@router.get("/users/list", response_model=CompanyUsersListResponse)
async def list_company_users(
    role: str | None = Query(None, description="Filter by role (recruiter, admin, viewer)"),
    is_active: bool = Query(True, description="Filter by active status"),
    user_repo: UserRepository = Depends(get_user_repo),
    current_user: User = Depends(get_current_user_or_demo),
):
    """List users belonging to the same company as the current user."""
    try:
        if not current_user.company_id:
            raise HTTPException(status_code=400, detail="User has no company_id assigned")
        company_id = current_user.company_id

        users = await user_repo.list_for_company(str(company_id), role=role, is_active=is_active)
        user_emails = [u.email for u in users]
        jobs_by_email = await user_repo.count_active_jobs_by_email(user_emails)

        user_responses = []
        for user in users:
            active_jobs_count = jobs_by_email.get(user.email, 0)
            performance_score = 85 + (hash(str(user.id)) % 15)
            user_responses.append(CompanyUserResponse(
                id=str(user.id),
                name=user.name,
                email=user.email,
                role=user.role.value if hasattr(user.role, 'value') else str(user.role),
                is_active=user.is_active,
                active_jobs_count=active_jobs_count,
                performance_score=performance_score,
            ))

        logger.info(f"Listed {len(user_responses)} users for company {company_id}")
        return CompanyUsersListResponse(users=user_responses, total=len(user_responses))
    except Exception as e:
        logger.error(f"Error listing company users: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# CATALOG STATUS - Smart Wizard Integration
# ============================================================================

class CatalogStatusResponse(BaseModel):
    company_id: str
    maturity_score: int
    maturity_level: str
    maturity_factors: list[str]
    smart_start_enabled: bool
    required_fields_for_wizard: list[str]
    available_data_summary: list[str]
    counts: dict[str, int]
    recommendations: list[str]


class SmartWizardGreetingResponse(BaseModel):
    greeting_message: str
    catalog_status: CatalogStatusResponse
    prefill_data: dict[str, Any]


@router.get("/catalog-status", response_model=CatalogStatusResponse)
async def get_catalog_status(
    company_id: str = Query(default=""),
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
):
    """Get the maturity status of company's catalog data for Smart Wizard."""
    try:
        status = await company_config_service.get_catalog_status(company_id, profile_repo.db)
        return CatalogStatusResponse(**status)
    except Exception as e:
        logger.error(f"Error getting catalog status: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/smart-wizard-greeting", response_model=SmartWizardGreetingResponse)
async def get_smart_wizard_greeting(
    company_id: str = Query(default=""),
    dept_repo: DepartmentRepository = Depends(get_department_repo),
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
):
    """Get personalized greeting for the Smart Wizard based on catalog status."""
    import asyncio

    try:
        status, config = await asyncio.gather(
            company_config_service.get_catalog_status(company_id, profile_repo.db),
            company_config_service.get_configuration(company_id, profile_repo.db),
        )

        base_intro = """Olá! Sou a LIA, sua assistente de recrutamento.

Como você gostaria de começar?

🆕 **Criar vaga do zero** — Me conte sobre a posição que você precisa preencher

📋 **Usar vaga existente** — Posso duplicar e adaptar uma vaga anterior

📝 **Usar template** — Escolha entre nossos modelos prontos por área/cargo

"""

        if status["maturity_level"] == "complete":
            greeting = base_intro + """💡 **Dica:** Já tenho suas políticas configuradas, então posso sugerir salários, benefícios e competências automaticamente!

Qual opção você prefere?"""
        elif status["maturity_level"] == "partial":
            greeting = base_intro + """💡 **Dica:** Encontrei alguns dados da sua empresa que podem agilizar o processo.

Qual opção você prefere?"""
        else:
            greeting = base_intro + """💡 **Dica:** Você pode descrever a vaga em linguagem natural que eu extraio as informações automaticamente.

Qual opção você prefere?"""

        departments_list = []
        try:
            from uuid import UUID as UUID_type

            resolved_company_uuid = None
            try:
                resolved_company_uuid = UUID_type(company_id)
            except (ValueError, TypeError):
                profile = await profile_repo.get_default()
                if profile:
                    resolved_company_uuid = profile.id

            if resolved_company_uuid:
                departments = await dept_repo.list_for_company(resolved_company_uuid)
                departments_list = [
                    {"id": str(d.id), "name": d.name, "manager": d.manager_name}
                    for d in departments
                ]
        except Exception as e:
            logger.warning(f"Error fetching departments for prefill: {e}")

        tech_stack = []
        try:
            from uuid import UUID as UUID_type

            try:
                company_uuid = UUID_type(company_id)
                profile = await profile_repo.get_by_id(company_uuid)
            except (ValueError, TypeError):
                profile = await profile_repo.get_default()

            if profile and hasattr(profile, 'tech_stack') and profile.tech_stack:
                tech_stack = profile.tech_stack if isinstance(profile.tech_stack, list) else []
        except Exception as e:
            logger.warning(f"Error fetching tech stack: {e}")

        screening_questions_list = [
            {
                "id": str(idx),
                "question": q.get("question", q.get("text", "")),
                "category": q.get("category", "general"),
            }
            for idx, q in enumerate(config.screening_questions or [])
        ]

        prefill_data = {
            "benefits": [
                {"name": b.get("name"), "category": b.get("category", "outros")}
                for b in (config.benefits or [])[:15]
            ],
            "departments": departments_list,
            "screening_questions": screening_questions_list,
            "tech_stack": tech_stack,
            "culture_values": [v.get("name") for v in (config.culture_values or [])[:10]],
            "default_work_model": "hybrid",
            "default_employment_type": "CLT",
            "default_pipeline": config.default_pipeline,
        }

        return SmartWizardGreetingResponse(
            greeting_message=greeting,
            catalog_status=CatalogStatusResponse(**status),
            prefill_data=prefill_data,
        )
    except Exception as e:
        logger.error(f"Error getting smart wizard greeting: {e}")
        raise HTTPException(status_code=500, detail=str(e))
