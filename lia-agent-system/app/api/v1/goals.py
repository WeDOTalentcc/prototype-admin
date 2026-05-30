"""
Goals API endpoints for managing user goals and targets.
"""
import csv
import io
import logging
import uuid
from datetime import datetime
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.domains.goals.dependencies import get_goals_repo
from app.domains.goals.repositories.goals_repository import GoalsRepository
from app.models.goal import Goal
from app.shared.security.require_company_id import require_company_id, require_company_id_strict_match
from app.shared.types import WeDoBaseModel
from typing import Annotated
from fastapi import Path
from app.api.v1._path_patterns import DUAL_ID_PATH_PATTERN, reorder_collection_before_item

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/goals", tags=["goals"])


class GoalCreate(WeDoBaseModel):
    user_id: str
    template_id: str | None = None
    name: str
    description: str | None = None
    target: float
    current: float = 0
    unit: str | None = None
    period: str = "monthly"
    category: str = "recruitment"
    start_date: datetime | None = None
    end_date: datetime | None = None
    is_custom: bool = False


class GoalUpdate(WeDoBaseModel):
    name: str | None = None
    description: str | None = None
    target: float | None = None
    current: float | None = None
    unit: str | None = None
    period: str | None = None
    category: str | None = None
    status: str | None = None
    start_date: datetime | None = None
    end_date: datetime | None = None
    is_active: bool | None = None


class GoalResponse(BaseModel):
    id: uuid.UUID
    user_id: str
    template_id: str | None
    name: str
    description: str | None
    target: float
    current: float
    unit: str | None
    period: str
    category: str
    status: str
    start_date: datetime | None
    end_date: datetime | None
    progress: float
    is_custom: bool
    is_active: bool
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class GoalTemplateCreate(WeDoBaseModel):
    name: str
    description: str | None = None
    category: str = "recruitment"
    default_target: float
    unit: str | None = None
    period: str = "monthly"


class GoalTemplateResponse(BaseModel):
    id: uuid.UUID
    name: str
    description: str | None
    category: str
    default_target: float
    unit: str | None
    period: str
    is_active: bool
    is_system: bool
    created_at: datetime

    class Config:
        from_attributes = True


@router.get("", response_model=list[GoalResponse])
async def list_goals(
    user_id: str | None = Query(None),
    company_id: uuid.UUID | None = Query(None),
    period: str | None = Query(None),
    category: str | None = Query(None),
    status: str | None = Query(None),
    include_inactive: bool = Query(False),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=200),
    repo: GoalsRepository = Depends(get_goals_repo),
_company_gate: str = Depends(require_company_id_strict_match("query.company_id"))):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """List all goals with optional filters."""
    try:
        goals = await repo.list_goals(
            user_id=user_id,
            company_id=company_id,
            period=period,
            category=category,
            status=status,
            include_inactive=include_inactive,
            skip=skip,
            limit=limit,
        )
        return goals
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing goals: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/by-user/{user_id}", response_model=None)
async def get_goals_by_user(
    user_id: Annotated[str, Path(pattern=DUAL_ID_PATH_PATTERN)],
    include_inactive: bool = Query(False),
    repo: GoalsRepository = Depends(get_goals_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Get all goals for a specific user, grouped by period."""
    try:
        goals = await repo.list_goals(user_id=user_id, include_inactive=include_inactive)

        grouped: dict[str, list[dict]] = {
            "monthly": [],
            "quarterly": [],
            "yearly": [],
        }

        for goal in goals:
            goal_data = {
                "id": str(goal.id),
                "userId": goal.user_id,
                "templateId": goal.template_id,
                "name": goal.name,
                "description": goal.description,
                "target": goal.target,
                "current": goal.current,
                "unit": goal.unit,
                "period": goal.period,
                "category": goal.category,
                "status": goal.status,
                "startDate": goal.start_date.isoformat() if goal.start_date else None,
                "endDate": goal.end_date.isoformat() if goal.end_date else None,
                "progress": goal.progress,
                "isCustom": goal.is_custom,
            }

            if goal.period in grouped:
                grouped[goal.period].append(goal_data)

        return grouped
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting goals for user {user_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{goal_id}", response_model=GoalResponse)
async def get_goal(
    goal_id: Annotated[str, Path(pattern=DUAL_ID_PATH_PATTERN)],
    repo: GoalsRepository = Depends(get_goals_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Get a specific goal by ID."""
    try:
        goal = await repo.get_by_id(goal_id)

        if not goal:
            raise HTTPException(status_code=404, detail="Goal not found")

        return goal
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching goal: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("", response_model=GoalResponse)
async def create_goal(
    data: GoalCreate,
    repo: GoalsRepository = Depends(get_goals_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Create a new goal."""
    try:
        goal_data = data.model_dump()

        if goal_data.get("start_date") and hasattr(goal_data["start_date"], "tzinfo") and goal_data["start_date"].tzinfo:
            goal_data["start_date"] = goal_data["start_date"].replace(tzinfo=None)
        if goal_data.get("end_date") and hasattr(goal_data["end_date"], "tzinfo") and goal_data["end_date"].tzinfo:
            goal_data["end_date"] = goal_data["end_date"].replace(tzinfo=None)

        goal = await repo.create(goal_data)

        # pii-logs ok: nome de entidade/config (não PII per LGPD Art.5 V — pessoa natural)
        logger.info(f"Created goal: {goal.name} for user {goal.user_id}")
        return goal
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating goal: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{goal_id}", response_model=GoalResponse)
async def update_goal(
    goal_id: Annotated[str, Path(pattern=DUAL_ID_PATH_PATTERN)],
    data: GoalUpdate,
    repo: GoalsRepository = Depends(get_goals_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Update an existing goal."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        goal = await repo.update(goal_id, update_data)

        if not goal:
            raise HTTPException(status_code=404, detail="Goal not found")

        # pii-logs ok: nome de entidade/config (não PII per LGPD Art.5 V — pessoa natural)
        logger.info(f"Updated goal: {goal.name}")
        return goal
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating goal: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{goal_id}", response_model=None)
async def delete_goal(
    goal_id: Annotated[str, Path(pattern=DUAL_ID_PATH_PATTERN)],
    repo: GoalsRepository = Depends(get_goals_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Soft delete a goal."""
    try:
        goal = await repo.soft_delete(goal_id)

        if not goal:
            raise HTTPException(status_code=404, detail="Goal not found")

        return {"success": True, "message": "Goal deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting goal: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/bulk", response_model=None)
async def create_goals_bulk(
    goals: list[GoalCreate],
    repo: GoalsRepository = Depends(get_goals_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Create multiple goals at once (for applying templates to multiple users)."""
    try:
        goals_data = [g.model_dump() for g in goals]
        created_goals = await repo.create_bulk(goals_data)

        logger.info(f"Created {len(created_goals)} goals in bulk")

        return {
            "success": True,
            "message": f"Created {len(created_goals)} goals",
            "goals": [GoalResponse.model_validate(g) for g in created_goals],
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating goals in bulk: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/templates/list", response_model=list[GoalTemplateResponse])
async def list_goal_templates(
    company_id: uuid.UUID | None = Query(None),
    category: str | None = Query(None),
    include_inactive: bool = Query(False),
    repo: GoalsRepository = Depends(get_goals_repo),
_company_gate: str = Depends(require_company_id_strict_match("query.company_id"))):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """List all goal templates."""
    try:
        templates = await repo.list_templates(
            company_id=company_id,
            category=category,
            include_inactive=include_inactive,
        )
        return templates
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing goal templates: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/templates", response_model=GoalTemplateResponse)
async def create_goal_template(
    data: GoalTemplateCreate,
    repo: GoalsRepository = Depends(get_goals_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Create a new goal template."""
    try:
        template = await repo.create_template(data.model_dump())

        # pii-logs ok: nome de entidade/config (não PII per LGPD Art.5 V — pessoa natural)
        logger.info(f"Created goal template: {template.name}")
        return template
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating goal template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/seed-templates", response_model=None)
async def seed_default_templates(
    repo: GoalsRepository = Depends(get_goals_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Seed default goal templates."""
    try:
        default_templates = [
            {
                "name": "Contratações Mensais",
                "description": "Número de contratações efetivadas no mês",
                "category": "recruitment",
                "default_target": 5,
                "unit": "contratações",
                "period": "monthly",
                "is_system": True,
            },
            {
                "name": "Time to Fill",
                "description": "Tempo médio para preenchimento de vagas em dias",
                "category": "efficiency",
                "default_target": 30,
                "unit": "dias",
                "period": "monthly",
                "is_system": True,
            },
            {
                "name": "NPS dos Candidatos",
                "description": "Score de satisfação dos candidatos no processo",
                "category": "satisfaction",
                "default_target": 85,
                "unit": "%",
                "period": "quarterly",
                "is_system": True,
            },
            {
                "name": "Taxa de Conversão",
                "description": "Percentual de candidatos convertidos em contratações",
                "category": "efficiency",
                "default_target": 2.5,
                "unit": "%",
                "period": "quarterly",
                "is_system": True,
            },
            {
                "name": "Score de Qualidade",
                "description": "Avaliação média da qualidade das contratações",
                "category": "quality",
                "default_target": 4.0,
                "unit": "pontos",
                "period": "yearly",
                "is_system": True,
            },
            {
                "name": "Entrevistas Mensais",
                "description": "Número de entrevistas realizadas no mês",
                "category": "recruitment",
                "default_target": 40,
                "unit": "entrevistas",
                "period": "monthly",
                "is_system": True,
            },
            {
                "name": "Taxa de Resposta",
                "description": "Percentual de candidatos que respondem ao contato inicial",
                "category": "efficiency",
                "default_target": 75,
                "unit": "%",
                "period": "monthly",
                "is_system": True,
            },
            {
                "name": "Taxa de Aceitação",
                "description": "Percentual de ofertas aceitas pelos candidatos",
                "category": "quality",
                "default_target": 85,
                "unit": "%",
                "period": "quarterly",
                "is_system": True,
            },
        ]

        created = 0
        skipped = 0

        for template_data in default_templates:
            existing = await repo.find_system_template_by_name(template_data["name"])
            if existing:
                skipped += 1
                continue

            await repo.create_template(template_data)
            created += 1

        return {
            "success": True,
            "message": f"Seeded {created} templates, skipped {skipped} existing",
            "created": created,
            "skipped": skipped,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error seeding goal templates: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class GoalImportResponse(BaseModel):
    success: bool
    imported_count: int
    error_count: int
    errors: list[dict[str, Any]]
    items: list[dict[str, Any]]


def parse_csv_file_goals(content: bytes) -> list[dict[str, str]]:
    """Parse CSV file content and return list of dictionaries."""
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def parse_excel_file_goals(content: bytes) -> list[dict[str, str]]:
    """Parse Excel file content and return list of dictionaries."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        result = []
        for row in rows[1:]:
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(cell) if cell is not None else ""
            result.append(row_dict)

        return result
    except ImportError:
        from fastapi import HTTPException

        raise HTTPException(
            status_code=400,
            detail="Excel file support requires openpyxl. Please upload a CSV file instead.",
        )


async def parse_goal_import_file(file: UploadFile) -> list[dict[str, str]]:
    """Parse uploaded file (CSV or Excel) and return data."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    content = await file.read()
    file_ext = file.filename.lower().split(".")[-1] if "." in file.filename else ""

    if file_ext == "csv":
        return parse_csv_file_goals(content)
    elif file_ext in ["xlsx", "xls"]:
        return parse_excel_file_goals(content)
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type: .{file_ext}. Allowed: .csv, .xlsx, .xls",
        )


@router.get("/import/template", response_model=None)
async def download_goals_import_template(company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Download CSV template for goals import."""
    try:
        headers = ["user_id", "name", "description", "target", "unit", "period", "category"]

        csv_content = ",".join(headers) + "\n"
        csv_content += "recruiter1,Monthly Hires,Number of hires per month,10,hires,monthly,recruitment\n"
        csv_content += "recruiter1,Time to Fill,Average days to fill positions,25,days,monthly,efficiency\n"
        csv_content += "recruiter2,Candidate NPS,Candidate satisfaction score,85,%,quarterly,satisfaction\n"

        buffer = io.BytesIO(csv_content.encode("utf-8"))
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=goals_import_template.csv"},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating goals import template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/import", response_model=GoalImportResponse)
async def import_goals(
    file: UploadFile = File(...),
    company_id: uuid.UUID | None = Query(None),
    repo: GoalsRepository = Depends(get_goals_repo),
_company_gate: str = Depends(require_company_id_strict_match("query.company_id"))):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """
    Import goals from Excel/CSV file with AI processing.
    Expected columns: user_id, name, target, period, category
    Returns list of created goals.
    """
    try:
        logger.info(f"Starting goals import from file: {file.filename}")

        rows = await parse_goal_import_file(file)

        if not rows:
            return GoalImportResponse(
                success=False,
                imported_count=0,
                error_count=0,
                errors=[{"message": "No data found in file"}],
                items=[],
            )

        imported_items: list[dict] = []
        errors: list[dict] = []

        valid_periods = ["monthly", "quarterly", "yearly"]
        valid_categories = ["recruitment", "efficiency", "quality", "satisfaction"]

        for idx, row in enumerate(rows, start=2):
            row_errors: list[str] = []

            user_id = row.get("user_id", "").strip()
            name = row.get("name", "").strip()
            target_str = row.get("target", "").strip()

            if not user_id:
                row_errors.append(f"Row {idx}: Missing required field 'user_id'")
            if not name:
                row_errors.append(f"Row {idx}: Missing required field 'name'")
            if not target_str:
                row_errors.append(f"Row {idx}: Missing required field 'target'")

            target = 0.0
            if target_str:
                try:
                    target = float(target_str)
                except ValueError:
                    row_errors.append(f"Row {idx}: 'target' must be a number")

            period = row.get("period", "").strip().lower() or "monthly"
            if period not in valid_periods:
                row_errors.append(
                    f"Row {idx}: Invalid period '{period}'. Must be one of: {', '.join(valid_periods)}"
                )

            category = row.get("category", "").strip().lower() or "recruitment"
            if category not in valid_categories:
                row_errors.append(
                    f"Row {idx}: Invalid category '{category}'. Must be one of: {', '.join(valid_categories)}"
                )

            if row_errors:
                errors.append({"row": idx, "data": row, "errors": row_errors})
                continue

            goal = Goal(
                user_id=user_id,
                company_id=company_id,
                name=name,
                description=row.get("description", "").strip() or None,
                target=target,
                current=0,
                unit=row.get("unit", "").strip() or None,
                period=period,
                category=category,
                status="pending",
                progress=0,
                is_custom=True,
                is_active=True,
            )

            try:
                await repo.add_and_flush(goal)

                imported_items.append(
                    {
                        "id": str(goal.id),
                        "user_id": goal.user_id,
                        "name": goal.name,
                        "target": goal.target,
                        "period": goal.period,
                        "category": goal.category,
                        "row": idx,
                    }
                )
            except Exception as flush_error:
                errors.append(
                    {
                        "row": idx,
                        "data": row,
                        "errors": [f"Row {idx}: Database error - {str(flush_error)}"],
                    }
                )

        if imported_items:
            await repo.commit()
            logger.info(f"Imported {len(imported_items)} goals successfully")

        return GoalImportResponse(
            success=len(errors) == 0,
            imported_count=len(imported_items),
            error_count=len(errors),
            errors=errors,
            items=imported_items,
        )
    except HTTPException:
        raise
    except Exception as e:
        await repo.rollback()
        logger.error(f"Error importing goals: {e}")
        raise HTTPException(status_code=500, detail=str(e))

reorder_collection_before_item(router)
