"""
Benefits Templates API endpoints.
Provides pre-registered benefit templates for recruiters.
"""
import csv
import io
import logging
from datetime import datetime
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db, get_tenant_db
from app.domains.company.repositories.benefit_template_repository import BenefitTemplateRepository
from app.domains.company.repositories.benefit_repository import BenefitRepository
from app.models.company import Benefit, BenefitTemplate
from app.shared.security.require_company_id import require_company_id, require_company_id_strict_match
from app.shared.errors import LIAError

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/benefits", tags=["benefits"])


class BenefitTemplateResponse(BaseModel):
    id: UUID
    name: str
    description: str | None
    category: str
    is_popular: bool
    is_active: bool
    order: int
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class BenefitTemplateListResponse(BaseModel):
    items: list[BenefitTemplateResponse]
    total: int
    by_category: dict


BENEFIT_TEMPLATES_DATA = [
    {"name": "Plano de Saúde Médico", "description": "Cobertura de consultas, exames e internações", "category": "health", "is_popular": True, "order": 1},
    {"name": "Plano Odontológico", "description": "Cobertura de procedimentos odontológicos", "category": "health", "is_popular": True, "order": 2},
    {"name": "Plano de Saúde Familiar", "description": "Extensão do plano para dependentes", "category": "health", "is_popular": True, "order": 3},
    {"name": "Seguro de Vida", "description": "Proteção financeira para a família em caso de morte", "category": "health", "is_popular": True, "order": 4},
    {"name": "Seguro Saúde Complementar", "description": "Cobertura adicional para procedimentos específicos", "category": "health", "is_popular": False, "order": 5},
    {"name": "Apoio Psicológico/Terapia", "description": "Sessões com profissionais de saúde mental", "category": "health", "is_popular": True, "order": 6},
    {"name": "Programas de Meditação e Mindfulness", "description": "Aplicativos e aulas para redução de estresse", "category": "health", "is_popular": False, "order": 7},
    {"name": "Cuidados com Saúde Mental", "description": "Programas específicos de bem-estar emocional", "category": "health", "is_popular": False, "order": 8},
    {"name": "Auxílio Funeral", "description": "Cobertura para despesas com funeral", "category": "health", "is_popular": False, "order": 9},
    {"name": "Auxílio Saúde Complementar", "description": "Subsídio para despesas médicas não cobertas", "category": "health", "is_popular": False, "order": 10},
    {"name": "Vale-Refeição", "description": "Cartão ou voucher para refeições prontas", "category": "food", "is_popular": True, "order": 1},
    {"name": "Vale-Alimentação", "description": "Cartão para compra de alimentos em supermercados", "category": "food", "is_popular": True, "order": 2},
    {"name": "Cesta Básica", "description": "Fornecimento periódico de alimentos essenciais", "category": "food", "is_popular": False, "order": 3},
    {"name": "Refeição no Local de Trabalho", "description": "Fornecimento de refeições pela empresa", "category": "food", "is_popular": False, "order": 4},
    {"name": "Café da Manhã Corporativo", "description": "Oferecimento de café e lanches matinais", "category": "food", "is_popular": False, "order": 5},
    {"name": "Benefícios Flexíveis de Alimentação", "description": "Cartões que permitem uso em múltiplos estabelecimentos", "category": "food", "is_popular": True, "order": 6},
    {"name": "Vale-Combustível", "description": "Subsídio para combustível de veículo próprio", "category": "transport", "is_popular": True, "order": 1},
    {"name": "Auxílio Mobilidade", "description": "Benefício para utilização de outros meios de transporte", "category": "transport", "is_popular": True, "order": 2},
    {"name": "Vouchers para Aplicativos de Transporte", "description": "Descontos em Uber, 99, etc.", "category": "transport", "is_popular": False, "order": 3},
    {"name": "Estacionamento Subsidiado", "description": "Custeio total ou parcial de estacionamento", "category": "transport", "is_popular": False, "order": 4},
    {"name": "Car Allowance", "description": "Valor em dinheiro para manutenção e uso de veículo próprio", "category": "transport", "is_popular": False, "order": 5},
    {"name": "Bicicleta Corporativa", "description": "Fornecimento de bicicletas para deslocamento", "category": "transport", "is_popular": False, "order": 6},
    {"name": "Auxílio Transporte Adicional", "description": "Complemento ao vale-transporte obrigatório", "category": "transport", "is_popular": False, "order": 7},
    {"name": "Auxílio-Educação", "description": "Subsídio para educação formal (graduação, pós-graduação)", "category": "education", "is_popular": True, "order": 1},
    {"name": "Reembolso de Cursos", "description": "Custeio de cursos de capacitação profissional", "category": "education", "is_popular": True, "order": 2},
    {"name": "Plataformas de Aprendizagem Online", "description": "Acesso a cursos e treinamentos digitais", "category": "education", "is_popular": True, "order": 3},
    {"name": "Educação Corporativa", "description": "Universidade ou centro de educação interno", "category": "education", "is_popular": False, "order": 4},
    {"name": "Treinamentos e Palestras", "description": "Programas de desenvolvimento profissional", "category": "education", "is_popular": False, "order": 5},
    {"name": "Auxílio para Idiomas", "description": "Custeio de aulas de línguas estrangeiras", "category": "education", "is_popular": True, "order": 6},
    {"name": "Bolsas de Estudo", "description": "Financiamento parcial ou total de estudos", "category": "education", "is_popular": False, "order": 7},
    {"name": "Programas de Mentoria", "description": "Acompanhamento profissional estruturado", "category": "education", "is_popular": False, "order": 8},
    {"name": "Certificações Profissionais", "description": "Custeio de certificações na área de atuação", "category": "education", "is_popular": False, "order": 9},
    {"name": "Auxílio Academia", "description": "Subsídio para academia de ginástica", "category": "quality_life", "is_popular": True, "order": 1},
    {"name": "Descontos em Academias", "description": "Parcerias com academias para descontos", "category": "quality_life", "is_popular": False, "order": 2},
    {"name": "Programas de Fitness Corporativo", "description": "Aulas de exercício no local de trabalho", "category": "quality_life", "is_popular": False, "order": 3},
    {"name": "Yoga e Pilates", "description": "Aulas oferecidas na empresa", "category": "quality_life", "is_popular": False, "order": 4},
    {"name": "Nutricionista Corporativo", "description": "Atendimento de nutricionista para colaboradores", "category": "quality_life", "is_popular": False, "order": 5},
    {"name": "Fisioterapia", "description": "Sessões de fisioterapia para colaboradores", "category": "quality_life", "is_popular": False, "order": 6},
    {"name": "Vale-Cultura", "description": "Subsídio para atividades culturais", "category": "quality_life", "is_popular": False, "order": 7},
    {"name": "Ingressos para Eventos", "description": "Acesso subsidiado a eventos culturais", "category": "quality_life", "is_popular": False, "order": 8},
    {"name": "Assinatura de Streaming", "description": "Custeio de plataformas de vídeo/música", "category": "quality_life", "is_popular": False, "order": 9},
    {"name": "Home Office", "description": "Possibilidade de trabalhar remotamente", "category": "quality_life", "is_popular": True, "order": 10},
    {"name": "Auxílio Home Office", "description": "Subsídio para estrutura de trabalho em casa", "category": "quality_life", "is_popular": True, "order": 11},
    {"name": "Horários Flexíveis", "description": "Flexibilidade para ajustar horário de trabalho", "category": "quality_life", "is_popular": True, "order": 12},
    {"name": "Participação nos Lucros e Resultados (PLR)", "description": "Bônus baseado em resultados", "category": "financial", "is_popular": True, "order": 1},
    {"name": "Comissão", "description": "Remuneração variável baseada em vendas", "category": "financial", "is_popular": False, "order": 2},
    {"name": "Previdência Privada", "description": "Complemento ao INSS para aposentadoria", "category": "financial", "is_popular": True, "order": 3},
    {"name": "Empréstimo Consignado", "description": "Crédito com desconto em folha", "category": "financial", "is_popular": False, "order": 4},
    {"name": "Programa de Indicação", "description": "Bônus por indicação de novos funcionários", "category": "financial", "is_popular": False, "order": 5},
    {"name": "Bônus de Desempenho", "description": "Recompensa por atingimento de metas", "category": "financial", "is_popular": True, "order": 6},
    {"name": "Décimo Quarto Salário", "description": "Gratificação adicional", "category": "financial", "is_popular": False, "order": 7},
    {"name": "Licença-Maternidade Estendida", "description": "Período maior que o obrigatório", "category": "family", "is_popular": True, "order": 1},
    {"name": "Licença-Paternidade Estendida", "description": "Período maior que o obrigatório", "category": "family", "is_popular": True, "order": 2},
    {"name": "Auxílio-Creche", "description": "Subsídio para creche ou babá", "category": "family", "is_popular": True, "order": 3},
    {"name": "Auxílio-Escolar", "description": "Subsídio para educação dos filhos", "category": "family", "is_popular": False, "order": 4},
    {"name": "Licença para Cuidado de Familiar", "description": "Afastamento para cuidar de parentes", "category": "family", "is_popular": False, "order": 5},
    {"name": "Licença Pet", "description": "Dias de folga para cuidar de animais de estimação", "category": "family", "is_popular": False, "order": 6},
    {"name": "Plano Veterinário", "description": "Cobertura de despesas veterinárias", "category": "family", "is_popular": False, "order": 7},
    {"name": "Auxílio Pet", "description": "Subsídio para cuidados com animais", "category": "family", "is_popular": False, "order": 8},
    {"name": "Pet-Friendly Office", "description": "Permissão para levar animais ao trabalho", "category": "family", "is_popular": False, "order": 9},
    {"name": "Seguro de Vida em Grupo", "description": "Proteção com prêmio reduzido", "category": "security", "is_popular": True, "order": 1},
    {"name": "Auxílio Óculos/Lentes de Contato", "description": "Subsídio para correção visual", "category": "security", "is_popular": False, "order": 2},
    {"name": "Auxílio Aparelho Auditivo", "description": "Cobertura para deficiência auditiva", "category": "security", "is_popular": False, "order": 3},
    {"name": "Programa de Vacinação", "description": "Campanhas de vacinação na empresa", "category": "security", "is_popular": False, "order": 4},
    {"name": "Check-up Anual", "description": "Exame de saúde periódico", "category": "security", "is_popular": True, "order": 5},
    {"name": "Desconto em Farmácias", "description": "Parcerias para medicamentos", "category": "security", "is_popular": False, "order": 6},
    {"name": "Desconto em Compras", "description": "Parcerias com lojas e estabelecimentos", "category": "security", "is_popular": False, "order": 7},
    {"name": "Notebook/Equipamento de Trabalho", "description": "Fornecimento de equipamentos", "category": "security", "is_popular": True, "order": 8},
    {"name": "Smartphone Corporativo", "description": "Fornecimento de celular", "category": "security", "is_popular": False, "order": 9},
    {"name": "Assinatura de Jornais e Revistas", "description": "Acesso a publicações", "category": "security", "is_popular": False, "order": 10},
    {"name": "Biblioteca Corporativa", "description": "Acervo de livros para empréstimo", "category": "security", "is_popular": False, "order": 11},
    {"name": "Espaço de Coworking", "description": "Acesso a espaços de trabalho compartilhado", "category": "security", "is_popular": False, "order": 12},
    {"name": "Congelamento de Óvulos", "description": "Custeio de procedimento de preservação de fertilidade", "category": "security", "is_popular": False, "order": 13},
    {"name": "Assistência à Fertilidade", "description": "Cobertura de tratamentos de reprodução assistida", "category": "security", "is_popular": False, "order": 14},
    {"name": "Programa de Diversidade e Inclusão", "description": "Iniciativas para ambiente inclusivo", "category": "security", "is_popular": False, "order": 15},
]


@router.get("/templates", response_model=BenefitTemplateListResponse)
async def get_benefit_templates(
    category: str | None = Query(None, description="Filter by category"),
    search: str | None = Query(None, description="Search by name"),
    popular_only: bool = Query(False, description="Show only popular templates"),
    db: AsyncSession = Depends(get_db), 
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """
    Get all available benefit templates.
    """
    try:
        repo = BenefitTemplateRepository(db)
        templates = await repo.list_templates(
            category=category,
            search=search,
            popular_only=popular_only,
        )

        by_category: dict = {}
        for template in templates:
            if template.category not in by_category:
                by_category[template.category] = 0
            by_category[template.category] += 1

        return BenefitTemplateListResponse(
            items=templates,  # type: ignore[arg-type]
            total=len(templates),
            by_category=by_category
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching benefit templates: {e}")
        raise LIAError(message="Erro interno do servidor")


@router.post("/seed-templates", response_model=None)
async def seed_benefit_templates(
    db: AsyncSession = Depends(get_tenant_db), 
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """
    Seed the database with pre-defined benefit templates.
    Only adds templates that don't already exist (by name).
    """
    try:
        repo = BenefitTemplateRepository(db)
        existing_count = await repo.count_all()

        if existing_count and existing_count > 0:
            return {
                "success": True,
                "message": f"Templates already exist ({existing_count} templates)",
                "created": 0,
                "total": existing_count
            }

        created_count = 0
        for template_data in BENEFIT_TEMPLATES_DATA:
            existing = await repo.get_by_name(template_data["name"])
            if not existing:
                template = BenefitTemplate(**template_data)
                await repo.create(template)
                created_count += 1

        logger.info(f"✅ Seeded {created_count} benefit templates")

        return {
            "success": True,
            "message": f"Successfully seeded {created_count} benefit templates",
            "created": created_count,
            "total": len(BENEFIT_TEMPLATES_DATA)
        }

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error seeding benefit templates: {e}")
        raise LIAError(message="Erro interno do servidor")


@router.get("/templates/{template_id}", response_model=BenefitTemplateResponse)
async def get_benefit_template(
    template_id: UUID,
    db: AsyncSession = Depends(get_db), 
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """
    Get a specific benefit template by ID.
    """
    try:
        repo = BenefitTemplateRepository(db)
        template = await repo.get_by_id(template_id)

        if not template:
            raise HTTPException(status_code=404, detail="Template not found")

        return template

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching benefit template: {e}")
        raise LIAError(message="Erro interno do servidor")


def parse_csv_file(content: bytes) -> list[dict[str, str]]:
    """Parse CSV file content and return list of dictionaries."""
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def parse_excel_file(content: bytes) -> list[dict[str, str]]:
    """Parse Excel file content and return list of dictionaries."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        result = []
        for row in rows[1:]:
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(cell) if cell is not None else ''
            result.append(row_dict)

        return result
    except ImportError:
        raise HTTPException(
            status_code=400,
            detail="Excel file support requires openpyxl. Please upload a CSV file instead."
        )


def parse_file_content(filename: str, content: bytes) -> list[dict[str, str]]:
    """Parse file content based on extension."""
    file_ext = filename.split('.')[-1].lower() if '.' in filename else ''

    if file_ext == 'csv':
        return parse_csv_file(content)
    elif file_ext in ['xlsx', 'xls']:
        return parse_excel_file(content)
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type: .{file_ext}. Allowed: .csv, .xlsx, .xls"
        )


def normalize_name(name: str) -> str:
    """Normalize benefit name for matching."""
    import unicodedata
    normalized = unicodedata.normalize('NFD', name.lower())
    normalized = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
    normalized = normalized.replace('-', ' ').replace('_', ' ')
    normalized = ' '.join(normalized.split())
    return normalized


def build_templates_lookup() -> dict[str, dict]:
    """Build a lookup dictionary for templates by normalized name."""
    lookup = {}
    for template in BENEFIT_TEMPLATES_DATA:
        normalized = normalize_name(template["name"])
        lookup[normalized] = template
    return lookup


class BenefitImportResponse(BaseModel):
    success: bool
    imported: int
    matched: int
    new: int
    errors: list[str]


@router.get("/import/template", response_model=None)
async def download_benefits_import_template(company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """
    Download CSV template for benefits import.
    """
    try:
        headers = [
            "name", "description", "category", "value_type", "value",
            "percentage_value", "value_details", "seniority_levels",
            "waiting_period_days", "is_mandatory", "is_highlighted",
            "is_discount", "provider"
        ]

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()

        for template in BENEFIT_TEMPLATES_DATA:
            row = {
                "name": template["name"],
                "description": template.get("description", ""),
                "category": template.get("category", ""),
                "value_type": "informative",
                "value": "",
                "percentage_value": "",
                "value_details": "",
                "seniority_levels": "",
                "waiting_period_days": "0",
                "is_mandatory": "false",
                "is_highlighted": str(template.get("is_popular", False)).lower(),
                "is_discount": "false",
                "provider": ""
            }
            writer.writerow(row)

        csv_content = output.getvalue()
        buffer = io.BytesIO(csv_content.encode('utf-8'))
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=template_beneficios.csv"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating benefits import template: {e}")
        raise LIAError(message="Erro interno do servidor")


@router.post("/import", response_model=BenefitImportResponse)
async def import_benefits(
    file: UploadFile = File(...),
    company_id: UUID = Query(..., description="Company ID to associate benefits with"),
    db: AsyncSession = Depends(get_tenant_db), 
_company_gate: str = Depends(require_company_id_strict_match("query.company_id"))):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """
    Import benefits from CSV or Excel file.
    """
    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Empty file uploaded")

        rows = parse_file_content(file.filename or "file.csv", content)
        benefit_repo = BenefitRepository(db)

        if not rows:
            raise HTTPException(status_code=400, detail="No data found in file")

        templates_lookup = build_templates_lookup()

        imported_count = 0
        matched_count = 0
        new_count = 0
        errors = []

        for idx, row in enumerate(rows, start=2):
            try:
                name = row.get('name', '').strip()
                if not name:
                    errors.append(f"Row {idx}: Missing required field 'name'")
                    continue

                normalized_name = normalize_name(name)
                template_match = templates_lookup.get(normalized_name)

                description = row.get('description', '').strip()
                category = row.get('category', '').strip()

                if template_match:
                    matched_count += 1
                    if not description:
                        description = template_match.get("description", "")
                    if not category:
                        category = template_match.get("category", "other")
                else:
                    new_count += 1
                    if not category:
                        category = "other"

                value_type = row.get('value_type', 'informative').strip().lower()
                if value_type not in ['monetary', 'percentage', 'informative']:
                    value_type = 'informative'

                value = None
                value_str = row.get('value', '').strip()
                if value_str:
                    try:
                        value = float(value_str.replace(',', '.'))
                    except ValueError:
                        pass

                percentage_value = None
                percentage_str = row.get('percentage_value', '').strip()
                if percentage_str:
                    try:
                        percentage_value = float(percentage_str.replace(',', '.').replace('%', ''))
                    except ValueError:
                        pass

                waiting_period = 0
                waiting_str = row.get('waiting_period_days', '').strip()
                if waiting_str:
                    try:
                        waiting_period = int(waiting_str)
                    except ValueError:
                        pass

                def parse_bool(val: str) -> bool:
                    return val.strip().lower() in ['true', '1', 'yes', 'sim', 's', 'x']

                seniority_levels_str = row.get('seniority_levels', '').strip()
                seniority_levels = []
                if seniority_levels_str:
                    seniority_levels = [s.strip() for s in seniority_levels_str.split(',') if s.strip()]

                benefit = Benefit(
                    company_id=company_id,
                    name=name,
                    description=description,
                    category=category,
                    value_type=value_type,
                    value=value,
                    percentage_value=percentage_value,
                    value_details=row.get('value_details', '').strip() or None,
                    seniority_levels=seniority_levels,
                    waiting_period_days=waiting_period,
                    is_mandatory=parse_bool(row.get('is_mandatory', '')),
                    is_highlighted=parse_bool(row.get('is_highlighted', '')),
                    is_discount=parse_bool(row.get('is_discount', '')),
                    provider=row.get('provider', '').strip() or None,
                    is_active=True
                )

                benefit_repo.db.add(benefit)
                imported_count += 1

            except Exception as row_error:
                errors.append(f"Row {idx}: {str(row_error)}")

        if imported_count > 0:
            logger.info(f"✅ Imported {imported_count} benefits ({matched_count} matched templates, {new_count} new)")

        return BenefitImportResponse(
            success=imported_count > 0,
            imported=imported_count,
            matched=matched_count,
            new=new_count,
            errors=errors
        )

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error importing benefits: {e}")
        raise LIAError(message="Erro interno do servidor")
