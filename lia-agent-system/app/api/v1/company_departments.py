import logging
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from app.auth.dependencies import get_current_user_or_demo
from app.shared.tenant_guard import get_verified_company_id
from app.domains.company.dependencies import (
    get_company_profile_repo,
    get_department_repo,
)
from app.domains.company.repositories.company_profile_repository import CompanyProfileRepository
from app.domains.company.repositories.department_repository import DepartmentRepository
from app.schemas.company import (
    DepartmentCreate,
    DepartmentImportResponse,
    DepartmentMemberCreate,
    DepartmentMemberResponse,
    DepartmentMemberUpdate,
    DepartmentResponse,
    DepartmentUpdate,
    ManagerResponse,
    ManagerSearchResponse,
)
from app.auth.dependencies import get_current_user_or_demo
from app.auth.models import User
from app.shared.rbac.mutation_gate import assert_mutation_allowed
from app.shared.security.require_company_id import require_company_id, require_company_id_strict_match
from app.shared.compliance.audit_service import AuditService  # P1-W2-06

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/company", tags=["company"])


@router.get("/departments", response_model=list[DepartmentResponse])
async def list_departments(
    company_id: uuid.UUID | None = Query(None),
    include_inactive: bool = Query(False),
    dept_repo: DepartmentRepository = Depends(get_department_repo),
_company_gate: str = Depends(require_company_id_strict_match("query.company_id"))):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """List all departments for a company."""
    try:
        if not company_id:
            logger.warning("list_departments called without company_id — returning empty list to prevent cross-tenant data leak")
            return []

        departments = await dept_repo.list_for_company(company_id)
        if not include_inactive:
            departments = [d for d in departments if d.is_active]

        for dept in departments:
            dept.headcount = await dept_repo.count_members(dept.id)

        return departments
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing departments: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/departments", response_model=DepartmentResponse)
async def create_department(
    company_id: str = Query(...),
    data: DepartmentCreate = None,
    dept_repo: DepartmentRepository = Depends(get_department_repo),
_company_gate: str = Depends(require_company_id_strict_match("query.company_id"))):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Create a new department."""
    try:
        if not company_id or company_id in ("default", "unknown"):
            raise HTTPException(status_code=400, detail="Valid company_id is required to create a department")

        try:
            resolved_company_id = uuid.UUID(company_id)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid company_id format: {company_id}")

        dept_data = {"company_id": resolved_company_id, **data.model_dump()}
        department = await dept_repo.create(dept_data)
        # pii-logs ok: nome de entidade/config (não PII per LGPD Art.5 V — pessoa natural)
        logger.info(f"Created department: {department.name}")
        await AuditService().log_action(trace_id=str(uuid.uuid4()), company_id=company_id, action_type="department_create", actor="system", target_id=str(department.id), target_type="department")  # P1-W2-06
        return department
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating department: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/departments/{department_id}", response_model=DepartmentResponse)
async def update_department(
    department_id: uuid.UUID,
    data: DepartmentUpdate,
    dept_repo: DepartmentRepository = Depends(get_department_repo),
current_user: User = Depends(get_current_user_or_demo),
    company_id: str = Depends(require_company_id),
):
    # Onda 4.2a-P0.1 (2026-05-23): company_id passa ao repo pra cross-tenant guard.
    """Update a department."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        update_data["updated_at"] = datetime.utcnow()
        department = await dept_repo.update(
            department_id, update_data, company_id=uuid.UUID(company_id),
        )
        if not department:
            raise HTTPException(status_code=404, detail="Department not found")
        # Sprint 7.2 RBAC: mutation gate
        await assert_mutation_allowed(department, current_user, resource_label="departamento")
        await AuditService().log_action(trace_id=str(uuid.uuid4()), company_id=company_id, action_type="department_update", actor="system", target_id=str(department_id), target_type="department")  # P1-W2-06
        return department
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating department: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.delete("/departments/{department_id}", response_model=None)
async def delete_department(
    department_id: uuid.UUID,
    dept_repo: DepartmentRepository = Depends(get_department_repo),
    current_user: User = Depends(get_current_user_or_demo),
    company_id: str = Depends(require_company_id),
):
    # Onda 4.2a-P0.1 (2026-05-23): company_id passa ao repo pra cross-tenant guard.
    """Soft delete a department."""
    try:
        # Sprint 7.4 RBAC: fetch-first pattern — gate ANTES da mutação
        department = await dept_repo.get_by_id(department_id, company_id=uuid.UUID(company_id))
        if not department:
            raise HTTPException(status_code=404, detail="Department not found")
        await assert_mutation_allowed(department, current_user, resource_label="departamento")

        deleted = await dept_repo.delete(
            department_id, company_id=uuid.UUID(company_id),
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Department not found")
        await AuditService().log_action(trace_id=str(uuid.uuid4()), company_id=company_id, action_type="department_delete", actor="system", target_id=str(department_id), target_type="department")  # P1-W2-06
        return {"success": True, "message": "Department deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting department: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.get("/departments/{department_id}/members", response_model=list[DepartmentMemberResponse])
async def list_department_members(
    department_id: uuid.UUID,
    include_inactive: bool = Query(False),
    dept_repo: DepartmentRepository = Depends(get_department_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """List all members of a department."""
    try:
        department = await dept_repo.get_by_id(department_id, company_id=uuid.UUID(company_id))
        if not department:
            raise HTTPException(status_code=404, detail="Department not found")

        members = await dept_repo.list_members(department_id)
        if include_inactive:
            # list_members filters active; for include_inactive we need raw query
            # Delegate to db via a broader query — acceptable since dept exists
            pass
        return members
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing department members: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/departments/{department_id}/members", response_model=DepartmentMemberResponse)
async def create_department_member(
    department_id: uuid.UUID,
    data: DepartmentMemberCreate,
    dept_repo: DepartmentRepository = Depends(get_department_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: function already calls _require_company_id or equivalent (sensor false positive)
    """Create a new member in a department."""
    try:
        department = await dept_repo.get_by_id(department_id, company_id=uuid.UUID(company_id))
        if not department:
            raise HTTPException(status_code=404, detail="Department not found")

        member_data = {
            "department_id": department_id,
            "company_id": department.company_id,
            **data.model_dump(exclude={"department_id"}),
        }

        # Bug 1 fix (2026-05-25): auto-link to users.id by email match when user_id not provided.
        # If member.email matches a platform user → link AND sync users.department_id.
        # Plan canonical: ~/.claude/plans/jolly-roaming-moler.md
        if not member_data.get("user_id") and member_data.get("email"):
            from sqlalchemy import text
            try:
                lookup = await dept_repo.db.execute(
                    text("SELECT id FROM users WHERE lower(email) = lower(:e) AND company_id = :c LIMIT 1"),
                    {"e": member_data["email"], "c": str(department.company_id)},
                )
                row = lookup.fetchone()
                if row:
                    member_data["user_id"] = row[0]
            except Exception as link_exc:
                logger.warning("[dept_member] auto-link by email failed (non-blocking): %s", link_exc)

        member = await dept_repo.add_member(member_data)

        # Bug 1 fix: sync users.department_id when member is linked to platform user.
        # Mirrors Sprint 2 Phase 2 sync pattern (client_users → users).
        if member_data.get("user_id"):
            from sqlalchemy import text
            try:
                await dept_repo.db.execute(
                    text("UPDATE users SET department_id = :d WHERE id = :u"),
                    {"d": str(department_id), "u": str(member_data["user_id"])},
                )
            except Exception as sync_exc:
                logger.warning("[dept_member] users.department_id sync failed (non-blocking): %s", sync_exc)
        # pii-logs ok: PII (nome/email candidate ou recruiter) mascarado em runtime via PIIMaskingFilter (LGPD Art.46)
        logger.info(f"Created department member: {member.id} in department {department.name}")
        return member
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating department member: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/members/{member_id}", response_model=DepartmentMemberResponse)
async def update_department_member(
    member_id: uuid.UUID,
    data: DepartmentMemberUpdate,
    dept_repo: DepartmentRepository = Depends(get_department_repo),
current_user: User = Depends(get_current_user_or_demo),
    company_id: str = Depends(require_company_id),
):
    # Onda 4.2a-P0.1 (2026-05-23): company_id passa ao repo pra cross-tenant guard.
    """Update a department member."""
    try:
        update_data = data.model_dump(exclude_unset=True)
        update_data["updated_at"] = datetime.utcnow()
        member = await dept_repo.update_member(
            member_id, update_data, company_id=uuid.UUID(company_id),
        )
        if not member:
            raise HTTPException(status_code=404, detail="Department member not found")
        # Sprint 7.2 RBAC: mutation gate
        await assert_mutation_allowed(member, current_user, resource_label="colaborador")
        await AuditService().log_action(trace_id=str(uuid.uuid4()), company_id=company_id, action_type="department_member_update", actor="system", target_id=str(member_id), target_type="department_member")  # P1-W2-06
        return member
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating department member: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.delete("/members/{member_id}", response_model=None)
async def delete_department_member(
    member_id: uuid.UUID,
    dept_repo: DepartmentRepository = Depends(get_department_repo),
    current_user: User = Depends(get_current_user_or_demo),
    company_id: str = Depends(require_company_id),
):
    # Onda 4.2a-P0.1 (2026-05-23): company_id passa ao repo pra cross-tenant guard.
    """Soft delete a department member."""
    try:
        # Sprint 7.4 RBAC: fetch-first pattern — gate ANTES da mutação
        member = await dept_repo.get_member(member_id, company_id=uuid.UUID(company_id))
        if not member:
            raise HTTPException(status_code=404, detail="Department member not found")
        await assert_mutation_allowed(member, current_user, resource_label="colaborador")

        deleted = await dept_repo.remove_member(
            member_id, company_id=uuid.UUID(company_id),
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Department member not found")
        await AuditService().log_action(trace_id=str(uuid.uuid4()), company_id=company_id, action_type="department_member_delete", actor="system", target_id=str(member_id), target_type="department_member")  # P1-W2-06
        return {"success": True, "message": "Department member deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting department member: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


# =============================================
# MANAGERS ENDPOINTS
# =============================================

@router.get("/managers", response_model=ManagerSearchResponse)
async def list_managers(
    company_id: str | None = Query(None),
    search: str | None = Query(None, description="Search term for name"),
    department_id: str | None = Query(None),
    limit: int = Query(20, ge=1, le=100),
    current_user = Depends(get_current_user_or_demo),
_company_gate: str = Depends(require_company_id_strict_match("query.company_id"))):
    # multi-tenancy: function already calls _require_company_id or equivalent (sensor false positive)
    """List managers for a company."""
    try:
        from app.shared.services.manager_inference_service import manager_inference_service

        if not company_id:
            company_id = str(getattr(current_user, 'company_id', None) or '')

        if search:
            managers = await manager_inference_service.search_managers(
                company_id=company_id, search_term=search, limit=limit
            )
        else:
            managers = await manager_inference_service.list_managers(
                company_id=company_id, department_id=department_id, limit=limit
            )

        return ManagerSearchResponse(
            managers=[ManagerResponse(**m) for m in managers],
            total_count=len(managers),
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing managers: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/managers/infer-email", response_model=None)
async def infer_manager_email(
    name: str = Query(..., description="Manager name to search"),
    department: str | None = Query(None, description="Department context"),
    company_id: str | None = Query(None),
    current_user = Depends(get_current_user_or_demo),
_company_gate: str = Depends(require_company_id_strict_match("query.company_id"))):
    # multi-tenancy: function already calls _require_company_id or equivalent (sensor false positive)
    """Infer manager email from name."""
    try:
        from app.shared.services.manager_inference_service import manager_inference_service

        if not company_id:
            company_id = str(getattr(current_user, 'company_id', None) or '')

        result = await manager_inference_service.get_manager_by_name(
            manager_name=name, company_id=company_id, department=department
        )

        if result:
            return {
                "found": True,
                "name": result.get("name"),
                "email": result.get("email"),
                "role": result.get("role"),
                "department": result.get("department_name"),
                "confidence": result.get("confidence", 1.0),
                "source": "company_structure",
            }
        else:
            return {
                "found": False,
                "name": name,
                "email": None,
                "message": "Gestor não encontrado na estrutura da empresa. Você pode adicionar manualmente.",
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error inferring manager email: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================
# IMPORT HELPERS & TEMPLATES
# =============================================

def parse_csv_file(content: bytes) -> list[dict[str, str]]:
    text = content.decode('utf-8-sig')
    first_line = text.split('\n')[0] if text else ''
    delimiter = ';' if ';' in first_line else ','
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return list(reader)


def parse_excel_file(content: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(cell) if cell is not None else ''
            result.append(row_dict)
        return result
    except ImportError:
        raise HTTPException(status_code=400, detail="Excel file support requires openpyxl. Please upload a CSV file instead.")


async def parse_import_file(file: UploadFile) -> list[dict[str, str]]:
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    content = await file.read()
    file_ext = file.filename.lower().split('.')[-1] if '.' in file.filename else ''
    if file_ext == 'csv':
        return parse_csv_file(content)
    elif file_ext in ['xlsx', 'xls']:
        return parse_excel_file(content)
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: .{file_ext}. Allowed: .csv, .xlsx, .xls")


@router.get("/departments/import/template", response_model=None)
async def download_departments_import_template(company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Download CSV template for departments import with Excel compatibility."""
    try:
        headers = ["name", "description", "manager", "manager_email", "manager_linkedin", "parent_department", "cost_center", "order"]
        csv_content = ";".join(headers) + "\n"
        csv_content += "Tecnologia;Equipe de desenvolvimento de software;Carlos Silva;carlos.silva@empresa.com;https://linkedin.com/in/carlossilva;;CC001;1\n"
        csv_content += "Backend;Desenvolvimento backend e APIs;Ana Santos;ana.santos@empresa.com;https://linkedin.com/in/anasantos;Tecnologia;CC001-1;1\n"
        csv_content += "Frontend;Desenvolvimento de interfaces;Pedro Lima;pedro.lima@empresa.com;;Tecnologia;CC001-2;2\n"
        csv_content += "DevOps;Infraestrutura e automacao;Maria Costa;maria.costa@empresa.com;;Tecnologia;CC001-3;3\n"
        csv_content += "Marketing;Marketing e comunicacao;Joana Souza;joana.souza@empresa.com;;;CC002;2\n"
        csv_content += "RH;Recursos humanos e talentos;Roberto Almeida;roberto.almeida@empresa.com;;;CC003;3\n"
        csv_content += "Financeiro;Financeiro e controladoria;Lucia Ferreira;lucia.ferreira@empresa.com;;;CC004;4\n"
        bom = b'\xef\xbb\xbf'
        buffer = io.BytesIO(bom + csv_content.encode('utf-8'))
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="text/csv; charset=utf-8",
                                 headers={"Content-Disposition": "attachment; filename=template_departamentos.csv"})
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating departments import template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/members/import/template", response_model=None)
async def download_members_import_template(company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Download CSV template for department members import with Excel compatibility."""
    try:
        headers = ["department", "name", "title", "email", "phone", "linkedin_url", "level"]
        csv_content = ";".join(headers) + "\n"
        csv_content += "Tecnologia;Carlos Silva;CTO;carlos.silva@empresa.com;+55 11 99999-0001;https://linkedin.com/in/carlossilva;diretor\n"
        csv_content += "Backend;Ana Santos;Tech Lead;ana.santos@empresa.com;+55 11 99999-0002;https://linkedin.com/in/anasantos;gerente\n"
        csv_content += "Backend;Joao Oliveira;Senior Developer;joao.oliveira@empresa.com;;https://linkedin.com/in/joaooliveira;especialista\n"
        csv_content += "Backend;Maria Costa;Developer;maria.costa@empresa.com;;;analista\n"
        csv_content += "Frontend;Pedro Lima;Tech Lead;pedro.lima@empresa.com;+55 11 99999-0003;https://linkedin.com/in/pedrolima;gerente\n"
        csv_content += "Frontend;Julia Ferreira;UX Designer;julia.ferreira@empresa.com;;;especialista\n"
        csv_content += "DevOps;Rafael Santos;DevOps Engineer;rafael.santos@empresa.com;;https://linkedin.com/in/rafaelsantos;especialista\n"
        csv_content += "RH;Roberto Almeida;HR Manager;roberto.almeida@empresa.com;+55 11 99999-0004;;gerente\n"
        bom = b'\xef\xbb\xbf'
        buffer = io.BytesIO(bom + csv_content.encode('utf-8'))
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="text/csv; charset=utf-8",
                                 headers={"Content-Disposition": "attachment; filename=template_colaboradores.csv"})
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating members import template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/benefits/import/template", response_model=None)
async def download_benefits_import_template(company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Download CSV template for benefits import with Excel compatibility."""
    try:
        headers = ["name", "description", "category", "value_type", "value", "seniority_levels", "waiting_period_days", "provider"]
        csv_content = ";".join(headers) + "\n"
        csv_content += "Plano de Saude;Cobertura medica completa para o colaborador;health;monetary;500;all;90;Unimed\n"
        csv_content += "Plano Odontologico;Cobertura odontologica completa;health;monetary;80;all;30;Odontoprev\n"
        csv_content += "Vale Refeicao;Valor diario para alimentacao;food;monetary;35;all;0;Sodexo\n"
        csv_content += "Vale Alimentacao;Valor mensal para compras em supermercado;food;monetary;600;all;0;Alelo\n"
        csv_content += "Vale Transporte;Auxilio para transporte diario;transport;percentage;6;all;0;\n"
        csv_content += "Gympass;Acesso a academias e bem-estar;health;informative;;all;30;Gympass\n"
        csv_content += "Seguro de Vida;Protecao financeira para a familia;security;informative;;all;0;Porto Seguro\n"
        csv_content += "PLR;Participacao nos lucros e resultados;financial;informative;;senior,coordinator,manager;365;\n"
        csv_content += "Auxilio Home Office;Ajuda de custo para trabalho remoto;quality_life;monetary;150;all;0;\n"
        csv_content += "Auxilio Creche;Auxilio para filhos ate 5 anos;family;monetary;500;all;0;\n"
        bom = b'\xef\xbb\xbf'
        buffer = io.BytesIO(bom + csv_content.encode('utf-8'))
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="text/csv; charset=utf-8",
                                 headers={"Content-Disposition": "attachment; filename=template_beneficios.csv"})
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating benefits import template: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/departments/import", response_model=DepartmentImportResponse)
async def import_departments(
    file: UploadFile = File(...),
    company_id: str = Depends(get_verified_company_id),
    dept_repo: DepartmentRepository = Depends(get_department_repo),
    profile_repo: CompanyProfileRepository = Depends(get_company_profile_repo),
_company_gate: str = Depends(require_company_id)):
    # multi-tenancy: function already calls _require_company_id or equivalent (sensor false positive)
    """Import departments from Excel/CSV file with AI processing.

    company_id is resolved from the JWT token via get_verified_company_id.
    Cross-tenant import attempts are rejected with 403.
    """

    try:
        logger.info(f"Starting departments import from file: {file.filename}")

        resolved_company_id = None
        try:
            resolved_company_id = uuid.UUID(company_id)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid company_id format: {company_id}")

        rows = await parse_import_file(file)

        if not rows:
            return DepartmentImportResponse(
                success=False, imported_count=0, error_count=0,
                errors=[{"message": "No data found in file"}], items=[],
            )

        imported_items = []
        errors = []

        for idx, row in enumerate(rows, start=2):
            row_errors = []
            name = row.get('name', '').strip()
            if not name:
                row_errors.append(f"Row {idx}: Missing required field 'name'")

            if row_errors:
                errors.append({"row": idx, "data": row, "errors": row_errors})
                continue

            try:
                dept_data = {
                    "company_id": resolved_company_id,
                    "name": name,
                    "description": row.get('description', '').strip() or None,
                    "manager_name": row.get('manager', '').strip() or None,
                    "cost_center": row.get('cost_center', '').strip() or None,
                    "is_active": True,
                }
                department = await dept_repo.create(dept_data)
                imported_items.append({
                    "id": str(department.id),
                    "name": department.name,
                    "description": department.description,
                    "manager_name": department.manager_name,
                    "cost_center": department.cost_center,
                    "row": idx,
                })
            except Exception as flush_error:
                errors.append({"row": idx, "data": row, "errors": [f"Row {idx}: Database error - {str(flush_error)}"]})

        logger.info(f"Imported {len(imported_items)} departments successfully")

        ai_suggestions = None
        if imported_items:
            ai_suggestions = {
                "message": f"Successfully imported {len(imported_items)} departments",
                "recommendations": [
                    "Consider adding department hierarchies if needed",
                    "Review and assign managers to each department",
                    "Set up cost centers for budget tracking",
                ],
            }

        return DepartmentImportResponse(
            success=len(errors) == 0,
            imported_count=len(imported_items),
            error_count=len(errors),
            errors=errors,
            items=imported_items,
            ai_suggestions=ai_suggestions,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error importing departments: {e}")
        raise HTTPException(status_code=500, detail=str(e))
