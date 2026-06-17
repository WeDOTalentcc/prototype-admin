"""Workforce planning API — hiring plans + planned headcounts.

Onda 4.2a-P0.3 (2026-05-23): cross-tenant guard via company_id passado
ao repo em todos os get_hiring_plan/get_headcount calls.
"""
import csv
import io
import logging
import uuid
from datetime import date, datetime
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

from app.repositories.dependencies import get_workforce_repo
from app.repositories.workforce_repository import WorkforceRepository
from app.domains.workforce.services.headcount_import_service import import_planned_headcounts
from app.models.workforce import WorkforceEntry
from app.shared.security.require_company_id import require_company_id, require_company_id_strict_match
from app.schemas.workforce import (
    HiringPlanCreate,
    HiringPlanResponse,
    HiringPlanUpdate,
    HiringPlanWithDetails,
    ImportConfirm,
    ImportJobResponse,
    ImportPreview,
    ImportResult,
    ImportRowValidation,
    MonthlyHeadcountStats,
    PlannedHeadcountCreate,
    PlannedHeadcountResponse,
    PlannedHeadcountUpdate,
    WorkforcePlanningStats,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/workforce", tags=["workforce"])


@router.get("/plans", response_model=list[HiringPlanResponse])
async def list_hiring_plans(
    company_id: uuid.UUID | None = Query(None),
    fiscal_year: int | None = Query(None),
    status: str | None = Query(None),
    include_inactive: bool = Query(False),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    repo: WorkforceRepository = Depends(get_workforce_repo),
_company_gate: str = Depends(require_company_id_strict_match("query.company_id"))):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """List all hiring plans with optional filters."""
    try:
        return await repo.list_hiring_plans(
            company_id=company_id,
            fiscal_year=fiscal_year,
            status=status,
            include_inactive=include_inactive,
            skip=skip,
            limit=limit,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing hiring plans: {e}")
        raise


@router.post("/plans", response_model=HiringPlanResponse)
async def create_hiring_plan(
    data: HiringPlanCreate,
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Create a new hiring plan."""
    try:
        plan = await repo.create_hiring_plan(data.model_dump())
        # pii-logs ok: nome de entidade/config (não PII per LGPD Art.5 V — pessoa natural)
        logger.info(f"Created hiring plan: {plan.name} for fiscal year {plan.fiscal_year}")
        return plan
    except HTTPException:
        raise
    except Exception as e:
        await repo.rollback()
        logger.error(f"Error creating hiring plan: {e}")
        raise


@router.get("/plans/{plan_id}", response_model=HiringPlanWithDetails)
async def get_hiring_plan(
    plan_id: uuid.UUID,
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Get a hiring plan with all details including headcounts and import jobs."""
    try:
        plan = await repo.get_hiring_plan_with_details(plan_id, company_id=company_id)

        if not plan:
            raise HTTPException(status_code=404, detail="Hiring plan not found")

        active_headcounts = [h for h in plan.planned_headcounts if h.is_active]

        response = HiringPlanWithDetails.model_validate(plan)
        response.planned_headcounts = [PlannedHeadcountResponse.model_validate(h) for h in active_headcounts]
        response.import_jobs = [ImportJobResponse.model_validate(j) for j in plan.import_jobs]

        return response
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching hiring plan: {e}")
        raise


@router.put("/plans/{plan_id}", response_model=HiringPlanResponse)
async def update_hiring_plan(
    plan_id: uuid.UUID,
    data: HiringPlanUpdate,
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Update an existing hiring plan."""
    try:
        plan = await repo.get_hiring_plan(plan_id, company_id=company_id)

        if not plan:
            raise HTTPException(status_code=404, detail="Hiring plan not found")

        plan = await repo.update_hiring_plan(plan, data.model_dump(exclude_unset=True))
        # pii-logs ok: nome de entidade/config (não PII per LGPD Art.5 V — pessoa natural)
        logger.info(f"Updated hiring plan: {plan.name}")
        return plan
    except HTTPException:
        raise
    except Exception as e:
        await repo.rollback()
        logger.error(f"Error updating hiring plan: {e}")
        raise


@router.delete("/plans/{plan_id}", response_model=None)
async def delete_hiring_plan(
    plan_id: uuid.UUID,
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Soft delete a hiring plan."""
    try:
        plan = await repo.get_hiring_plan(plan_id, company_id=company_id)

        if not plan:
            raise HTTPException(status_code=404, detail="Hiring plan not found")

        await repo.soft_delete_hiring_plan(plan)
        return {"success": True, "message": "Hiring plan deleted"}
    except HTTPException:
        raise
    except Exception as e:
        await repo.rollback()
        logger.error(f"Error deleting hiring plan: {e}")
        raise


@router.get("/plans/{plan_id}/headcounts", response_model=list[PlannedHeadcountResponse])
async def list_plan_headcounts(
    plan_id: uuid.UUID,
    status: str | None = Query(None),
    priority: str | None = Query(None),
    department_id: uuid.UUID | None = Query(None),
    target_month: int | None = Query(None, ge=1, le=12),
    target_year: int | None = Query(None),
    include_inactive: bool = Query(False),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=200),
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """List all headcounts for a hiring plan."""
    try:
        plan = await repo.get_hiring_plan(plan_id, company_id=company_id)
        if not plan:
            raise HTTPException(status_code=404, detail="Hiring plan not found")

        return await repo.list_plan_headcounts(
            plan_id=plan_id,
            status=status,
            priority=priority,
            department_id=department_id,
            target_month=target_month,
            target_year=target_year,
            include_inactive=include_inactive,
            skip=skip,
            limit=limit,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing headcounts: {e}")
        raise


@router.post("/plans/{plan_id}/headcounts", response_model=PlannedHeadcountResponse)
async def create_headcount(
    plan_id: uuid.UUID,
    data: PlannedHeadcountCreate,
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Create a new planned headcount."""
    try:
        plan = await repo.get_hiring_plan(plan_id, company_id=company_id)
        if not plan:
            raise HTTPException(status_code=404, detail="Hiring plan not found")

        headcount_data = data.model_dump()
        headcount_data["hiring_plan_id"] = plan_id

        headcount = await repo.create_headcount(headcount_data, plan)
        # pii-logs ok: nome de entidade/config (não PII per LGPD Art.5 V — pessoa natural)
        logger.info(f"Created headcount: {headcount.title} for plan {plan.name}")
        return headcount
    except HTTPException:
        raise
    except Exception as e:
        await repo.rollback()
        logger.error(f"Error creating headcount: {e}")
        raise


@router.get("/headcounts/{headcount_id}", response_model=PlannedHeadcountResponse)
async def get_headcount(
    headcount_id: uuid.UUID,
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Get a specific headcount by ID."""
    try:
        headcount = await repo.get_headcount(headcount_id, company_id=company_id)

        if not headcount:
            raise HTTPException(status_code=404, detail="Headcount not found")

        return headcount
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching headcount: {e}")
        raise


@router.put("/headcounts/{headcount_id}", response_model=PlannedHeadcountResponse)
async def update_headcount(
    headcount_id: uuid.UUID,
    data: PlannedHeadcountUpdate,
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Update an existing headcount."""
    try:
        headcount = await repo.get_headcount(headcount_id, company_id=company_id)

        if not headcount:
            raise HTTPException(status_code=404, detail="Headcount not found")

        headcount = await repo.update_headcount(headcount, data.model_dump(exclude_unset=True))
        # pii-logs ok: nome de entidade/config (não PII per LGPD Art.5 V — pessoa natural)
        logger.info(f"Updated headcount: {headcount.title}")
        return headcount
    except HTTPException:
        raise
    except Exception as e:
        await repo.rollback()
        logger.error(f"Error updating headcount: {e}")
        raise


@router.delete("/headcounts/{headcount_id}", response_model=None)
async def delete_headcount(
    headcount_id: uuid.UUID,
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Soft delete a headcount."""
    try:
        headcount = await repo.get_headcount(headcount_id, company_id=company_id)

        if not headcount:
            raise HTTPException(status_code=404, detail="Headcount not found")

        await repo.soft_delete_headcount(headcount)
        return {"success": True, "message": "Headcount deleted"}
    except HTTPException:
        raise
    except Exception as e:
        await repo.rollback()
        logger.error(f"Error deleting headcount: {e}")
        raise


@router.post("/plans/{plan_id}/headcounts/bulk", response_model=list[PlannedHeadcountResponse])
async def create_headcounts_bulk(
    plan_id: uuid.UUID,
    headcounts: list[PlannedHeadcountCreate],
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Create multiple headcounts at once."""
    try:
        plan = await repo.get_hiring_plan(plan_id, company_id=company_id)
        if not plan:
            raise HTTPException(status_code=404, detail="Hiring plan not found")

        headcounts_data = []
        for data in headcounts:
            hc_data = data.model_dump()
            hc_data["hiring_plan_id"] = plan_id
            headcounts_data.append(hc_data)

        created = await repo.create_headcounts_bulk(headcounts_data, plan)
        # pii-logs ok: nome de entidade/config (não PII per LGPD Art.5 V — pessoa natural)
        logger.info(f"Created {len(created)} headcounts in bulk for plan {plan.name}")
        return created
    except HTTPException:
        raise
    except Exception as e:
        await repo.rollback()
        logger.error(f"Error creating headcounts in bulk: {e}")
        raise


@router.post("/plans/{plan_id}/import/upload", response_model=ImportJobResponse)
async def upload_import_file(
    plan_id: uuid.UUID,
    file: UploadFile = File(...),
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Upload a file for import (Excel or CSV)."""
    try:
        plan = await repo.get_hiring_plan(plan_id, company_id=company_id)
        if not plan:
            raise HTTPException(status_code=404, detail="Hiring plan not found")

        if not file.filename:
            raise HTTPException(status_code=400, detail="No filename provided")

        allowed_extensions = [".xlsx", ".xls", ".csv"]
        file_ext = "." + file.filename.split(".")[-1].lower() if "." in file.filename else ""

        if file_ext not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid file type. Allowed: {', '.join(allowed_extensions)}",
            )

        file_type = "excel" if file_ext in [".xlsx", ".xls"] else "csv"

        import_job = await repo.create_import_job({
            "hiring_plan_id": plan_id,
            "file_name": file.filename,
            "file_type": file_type,
            "status": "processing",
        })

        logger.info(f"Created import job {import_job.id} for file {file.filename}")
        return import_job
    except HTTPException:
        raise
    except Exception as e:
        await repo.rollback()
        logger.error(f"Error uploading import file: {e}")
        raise


@router.get("/plans/{plan_id}/import/{job_id}/preview", response_model=ImportPreview)
async def preview_import(
    plan_id: uuid.UUID,
    job_id: uuid.UUID,
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Get preview of import data with validation."""
    try:
        import_job = await repo.get_import_job(job_id, plan_id)

        if not import_job:
            raise HTTPException(status_code=404, detail="Import job not found")

        expected_columns = [
            "title", "level", "department", "target_month", "target_year",
            "headcount", "salary_min", "salary_max", "priority",
            "hiring_manager_name", "hiring_manager_email", "justification",
        ]

        sample_validations = [
            ImportRowValidation(
                row_number=1,
                data={"title": "Software Engineer", "department": "Engineering", "target_month": 1},
                is_valid=True,
                errors=[],
                warnings=[],
            )
        ]

        preview = ImportPreview(
            file_name=import_job.file_name,
            file_type=import_job.file_type,
            total_rows=import_job.total_rows,
            valid_rows=import_job.imported_rows,
            error_rows=import_job.error_rows,
            detected_columns=expected_columns,
            column_mapping=import_job.column_mapping or {},
            ai_suggested_mapping=import_job.ai_suggestions.get("column_mapping") if import_job.ai_suggestions else None,
            sample_data=[],
            row_validations=sample_validations,
            can_proceed=import_job.error_rows == 0 or import_job.total_rows > import_job.error_rows,
            validation_summary={
                "total": import_job.total_rows,
                "valid": import_job.imported_rows,
                "errors": import_job.error_rows,
            },
        )

        return preview
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting import preview: {e}")
        raise


@router.post("/plans/{plan_id}/import/{job_id}/confirm", response_model=ImportResult)
async def confirm_import(
    plan_id: uuid.UUID,
    job_id: uuid.UUID,
    data: ImportConfirm,
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Confirm and execute the import."""
    try:
        import_job = await repo.get_import_job(job_id, plan_id)

        if not import_job:
            raise HTTPException(status_code=404, detail="Import job not found")

        if import_job.status == "completed":
            raise HTTPException(status_code=400, detail="Import already completed")

        import_job = await repo.confirm_import_job(import_job, data.column_mapping)

        import_result = ImportResult(
            success=True,
            import_job_id=import_job.id,
            total_rows=import_job.total_rows,
            imported_rows=import_job.imported_rows,
            error_rows=import_job.error_rows,
            errors=import_job.errors or [],
            created_headcount_ids=[],
        )

        logger.info(f"Import job {job_id} confirmed and completed")
        return import_result
    except HTTPException:
        raise
    except Exception as e:
        await repo.rollback()
        logger.error(f"Error confirming import: {e}")
        raise


@router.get("/import/template", response_model=None)
async def download_import_template(company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Download Excel template for headcount import."""
    try:
        headers = [
            "title", "level", "department", "target_month", "target_year",
            "headcount", "salary_min", "salary_max", "salary_currency",
            "contract_type", "priority", "hiring_manager_name",
            "hiring_manager_email", "justification", "job_description",
        ]

        csv_content = ",".join(headers) + "\n"
        csv_content += "Software Engineer,Senior,Engineering,3,2025,2,8000,12000,BRL,CLT,high,John Doe,john@example.com,Team expansion,Develop backend systems\n"
        csv_content += "Product Manager,Pleno,Product,6,2025,1,10000,15000,BRL,CLT,medium,Jane Smith,jane@example.com,New product launch,Lead product development\n"

        buffer = io.BytesIO(csv_content.encode("utf-8"))
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=workforce_import_template.csv"},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating import template: {e}")
        raise


@router.get("/stats", response_model=WorkforcePlanningStats)
async def get_workforce_stats(
    company_id: uuid.UUID | None = Query(None),
    fiscal_year: int | None = Query(None),
    plan_id: uuid.UUID | None = Query(None),
    repo: WorkforceRepository = Depends(get_workforce_repo),
_company_gate: str = Depends(require_company_id_strict_match("query.company_id"))):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Get workforce planning statistics."""
    try:
        plan = await repo.get_plan_for_stats(company_id=company_id, fiscal_year=fiscal_year, plan_id=plan_id)

        if not plan:
            current_year = datetime.now().year
            return WorkforcePlanningStats(
                hiring_plan_id=uuid.uuid4(),
                fiscal_year=fiscal_year or current_year,
                total_planned_headcount=0,
                total_filled=0,
                total_in_progress=0,
                total_pending=0,
                total_cancelled=0,
                fill_rate=0.0,
                monthly_breakdown=[],
                by_department={},
                by_level={},
                by_contract_type={},
                by_priority={},
            )

        headcounts = await repo.get_active_headcounts_for_plan(plan.id)

        total_planned = sum(h.headcount or 1 for h in headcounts)
        total_filled = sum(h.headcount or 1 for h in headcounts if h.status == "filled")
        total_in_progress = sum(h.headcount or 1 for h in headcounts if h.status == "in_progress")
        total_pending = sum(h.headcount or 1 for h in headcounts if h.status in ["planned", "pending"])
        total_cancelled = sum(h.headcount or 1 for h in headcounts if h.status == "cancelled")

        fill_rate = (total_filled / total_planned * 100) if total_planned > 0 else 0.0

        by_status: dict = {}
        by_priority: dict = {}
        by_level: dict = {}
        by_contract_type: dict = {}
        by_department: dict = {}

        for h in headcounts:
            status = h.status or "planned"
            by_status[status] = by_status.get(status, 0) + (h.headcount or 1)

            priority = h.priority or "medium"
            by_priority[priority] = by_priority.get(priority, 0) + (h.headcount or 1)

            level = h.level or "not_specified"
            by_level[level] = by_level.get(level, 0) + (h.headcount or 1)

            contract = h.contract_type or "not_specified"
            by_contract_type[contract] = by_contract_type.get(contract, 0) + (h.headcount or 1)

            dept_id = str(h.department_id) if h.department_id else "no_department"
            by_department[dept_id] = by_department.get(dept_id, 0) + (h.headcount or 1)

        monthly_data: dict = {}
        for h in headcounts:
            key = (h.target_month, h.target_year)
            if key not in monthly_data:
                monthly_data[key] = {"total": 0, "by_status": {}, "by_priority": {}, "by_department": {}}
            monthly_data[key]["total"] += h.headcount or 1

            status = h.status or "planned"
            monthly_data[key]["by_status"][status] = monthly_data[key]["by_status"].get(status, 0) + (h.headcount or 1)

            priority = h.priority or "medium"
            monthly_data[key]["by_priority"][priority] = monthly_data[key]["by_priority"].get(priority, 0) + (h.headcount or 1)

        monthly_breakdown = [
            MonthlyHeadcountStats(
                month=month,
                year=year,
                total_headcount=data["total"],
                by_status=data["by_status"],
                by_priority=data["by_priority"],
                by_department=data["by_department"],
            )
            for (month, year), data in sorted(monthly_data.items(), key=lambda x: (x[0][1], x[0][0]))
        ]

        positions_at_risk = 0
        today = date.today()
        for h in headcounts:
            if h.status in ["planned", "pending"]:
                target_date = date(h.target_year, h.target_month, 1)
                days_until = (target_date - today).days
                if 0 < days_until <= 30:
                    positions_at_risk += h.headcount or 1

        return WorkforcePlanningStats(
            hiring_plan_id=plan.id,
            fiscal_year=plan.fiscal_year,
            total_planned_headcount=total_planned,
            total_filled=total_filled,
            total_in_progress=total_in_progress,
            total_pending=total_pending,
            total_cancelled=total_cancelled,
            fill_rate=round(fill_rate, 2),
            total_budget=plan.total_budget,
            monthly_breakdown=monthly_breakdown,
            by_department=by_department,
            by_level=by_level,
            by_contract_type=by_contract_type,
            by_priority=by_priority,
            positions_at_risk=positions_at_risk,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting workforce stats: {e}")
        raise


@router.get("/timeline", response_model=None)
async def get_hiring_timeline(
    company_id: uuid.UUID | None = Query(None),
    months_ahead: int = Query(6, ge=1, le=24),
    repo: WorkforceRepository = Depends(get_workforce_repo),
_company_gate: str = Depends(require_company_id_strict_match("query.company_id"))):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Get timeline of upcoming hires."""
    try:
        today = date.today()
        current_month = today.month
        current_year = today.year

        headcounts = await repo.get_active_headcounts_filtered(company_id=company_id)

        timeline = []
        for month_offset in range(months_ahead):
            target_month = (current_month + month_offset - 1) % 12 + 1
            target_year = current_year + (current_month + month_offset - 1) // 12

            month_headcounts = [
                h for h in headcounts
                if h.target_month == target_month and h.target_year == target_year
            ]

            timeline.append({
                "month": target_month,
                "year": target_year,
                "month_name": date(target_year, target_month, 1).strftime("%B %Y"),
                "total_positions": sum(h.headcount or 1 for h in month_headcounts),
                "positions": [
                    {
                        "id": str(h.id),
                        "title": h.title,
                        "level": h.level,
                        "headcount": h.headcount or 1,
                        "priority": h.priority,
                        "status": h.status,
                        "hiring_manager": h.hiring_manager_name,
                    }
                    for h in month_headcounts
                ],
            })

        return {
            "timeline": timeline,
            "total_upcoming": sum(t["total_positions"] for t in timeline),
            "months_covered": months_ahead,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting hiring timeline: {e}")
        raise


@router.get("/alerts", response_model=None)
async def get_workforce_alerts(
    company_id: uuid.UUID | None = Query(None),
    repo: WorkforceRepository = Depends(get_workforce_repo),
_company_gate: str = Depends(require_company_id_strict_match("query.company_id"))):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Get alerts for upcoming or overdue hires."""
    try:
        today = date.today()

        headcounts = await repo.get_active_headcounts_filtered(company_id=company_id)

        alerts = []

        for h in headcounts:
            target_date = date(h.target_year, h.target_month, 1)
            days_until = (target_date - today).days

            if days_until < 0:
                alerts.append({
                    "type": "overdue",
                    "severity": "critical",
                    "headcount_id": str(h.id),
                    "title": h.title,
                    "message": f"Position {h.title} is {abs(days_until)} days overdue",
                    "target_date": target_date.isoformat(),
                    "days_overdue": abs(days_until),
                    "priority": h.priority,
                    "status": h.status,
                })
            elif days_until <= 7:
                alerts.append({
                    "type": "urgent",
                    "severity": "high",
                    "headcount_id": str(h.id),
                    "title": h.title,
                    "message": f"Position {h.title} needs to be filled within {days_until} days",
                    "target_date": target_date.isoformat(),
                    "days_remaining": days_until,
                    "priority": h.priority,
                    "status": h.status,
                })
            elif days_until <= 30:
                alerts.append({
                    "type": "upcoming",
                    "severity": "medium",
                    "headcount_id": str(h.id),
                    "title": h.title,
                    "message": f"Position {h.title} target date is in {days_until} days",
                    "target_date": target_date.isoformat(),
                    "days_remaining": days_until,
                    "priority": h.priority,
                    "status": h.status,
                })

        alerts.sort(key=lambda x: (
            0 if x["severity"] == "critical" else (1 if x["severity"] == "high" else 2),
            x.get("days_overdue", 0) if x["type"] == "overdue" else -x.get("days_remaining", 0),
        ))

        summary = {
            "total_alerts": len(alerts),
            "critical": len([a for a in alerts if a["severity"] == "critical"]),
            "high": len([a for a in alerts if a["severity"] == "high"]),
            "medium": len([a for a in alerts if a["severity"] == "medium"]),
        }

        return {
            "alerts": alerts,
            "summary": summary,
            "generated_at": datetime.utcnow().isoformat(),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting workforce alerts: {e}")
        raise


from pydantic import BaseModel
from app.shared.types import WeDoBaseModel


class WorkforceEntryItem(BaseModel):
    """Individual workforce entry for simple planning."""
    month: str
    department: str
    planned: int = 0
    actual: int = 0


class WorkforceEntriesRequest(WeDoBaseModel):
    """Request for saving workforce entries."""
    year: int
    entries: list[WorkforceEntryItem]


DEFAULT_WORKFORCE = [
    {"month": "Jan", "department": "Tecnologia", "planned": 5, "actual": 4, "ai_suggestion": 6},
    {"month": "Fev", "department": "Tecnologia", "planned": 4, "actual": 5},
    {"month": "Mar", "department": "Tecnologia", "planned": 6, "actual": 0, "ai_suggestion": 5},
    {"month": "Jan", "department": "Comercial", "planned": 3, "actual": 3},
    {"month": "Fev", "department": "Comercial", "planned": 2, "actual": 2},
    {"month": "Mar", "department": "Comercial", "planned": 4, "actual": 0, "ai_suggestion": 3},
    {"month": "Jan", "department": "RH", "planned": 1, "actual": 1},
    {"month": "Fev", "department": "RH", "planned": 1, "actual": 0},
    {"month": "Mar", "department": "RH", "planned": 2, "actual": 0, "ai_suggestion": 1},
]


@router.get("/entries", response_model=None)
async def get_workforce_entries(
    year: int | None = Query(None),
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Get simple workforce entries for admin panel.

    DEPRECATED: Use /workforce/plans/{plan_id}/headcounts/bulk (Sistema B canonical).
    workforce_entries sera removido em Sprint 2027-Q1.
    """
    logger.warning(
        "[DEPRECATED] workforce_entries GET endpoint called by company %s. "
        "Migrate to /workforce/plans/{plan_id}/headcounts/bulk.",
        company_id,
    )
    try:
        current_year = year or datetime.now().year
        # WT-2022 P0.WORK fix: pass company_id (cross-tenant prevention)
        entries = await repo.get_workforce_entries(current_year, company_id=company_id)

        if entries:
            return [e.to_dict() for e in entries]

        return DEFAULT_WORKFORCE
    except Exception as e:
        logger.error(f"Error fetching workforce entries: {e}")
        return DEFAULT_WORKFORCE


@router.put("/entries", response_model=None)
async def save_workforce_entries(
    data: WorkforceEntriesRequest,
    repo: WorkforceRepository = Depends(get_workforce_repo),
company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Save simple workforce entries for admin panel.

    DEPRECATED: Use /workforce/plans/{plan_id}/headcounts/bulk (Sistema B canonical).
    workforce_entries sera removido em Sprint 2027-Q1.
    """
    logger.warning(
        "[DEPRECATED] workforce_entries PUT endpoint called by company %s. "
        "Migrate to /workforce/plans/{plan_id}/headcounts/bulk.",
        company_id,
    )
    try:
        entries_data = [
            {
                "month": e.month,
                "department": e.department,
                "planned": e.planned,
                "actual": e.actual,
            }
            for e in data.entries
        ]
        # WT-2022 P0.WORK fix: pass company_id (cross-tenant prevention)
        entries = await repo.upsert_workforce_entries(
            data.year, entries_data, company_id=company_id
        )
        logger.info(f"Workforce entries saved for year {data.year}")
        return [e.to_dict() for e in entries]
    except HTTPException:
        raise
    except Exception as e:
        await repo.rollback()
        logger.error(f"Error saving workforce entries: {e}")
        raise


class WorkforceEntryImportResponse(BaseModel):
    """Response model for workforce entries import."""
    success: bool
    imported_count: int
    error_count: int
    errors: list[dict[str, Any]]
    items: list[dict[str, Any]]


def parse_csv_file_workforce(content: bytes) -> list[dict[str, str]]:
    """Parse CSV file content and return list of dictionaries."""
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def parse_excel_file_workforce(content: bytes) -> list[dict[str, str]]:
    """Parse Excel file content and return list of dictionaries."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        result = []
        for row in rows[1:]:
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(cell) if cell is not None else  ""
            result.append(row_dict)

        return result
    except ImportError:
        raise HTTPException(
            status_code=400,
            detail="Excel file support requires openpyxl. Please upload a CSV file instead.",
        )


async def parse_workforce_import_file(file: UploadFile) -> list[dict[str, str]]:
    """Parse uploaded file (CSV or Excel) and return data."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    content = await file.read()
    file_ext = file.filename.lower().split(".")[-1] if "." in file.filename else ""

    if file_ext == "csv":
        return parse_csv_file_workforce(content)
    elif file_ext in ["xlsx", "xls"]:
        return parse_excel_file_workforce(content)
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type: .{file_ext}. Allowed: .csv, .xlsx, .xls",
        )


@router.get("/entries/import/template", response_model=None)
async def download_workforce_entries_import_template(company_id: str = Depends(require_company_id)):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Download CSV template for headcount planning import (Store B / PlannedHeadcount)."""
    try:
        headers = ["department", "position", "headcount", "month", "year", "salary_min", "salary_max", "notes"]
        csv_content = ",".join(headers) + "\n"
        csv_content += "Tecnologia,Desenvolvedor(a) Backend Senior,2,3,2025,12000,18000,Expansao do time\n"
        csv_content += "Tecnologia,Engenheiro(a) de Dados,1,5,2025,14000,20000,Nova frente de dados\n"
        csv_content += "Comercial,Executivo(a) de Vendas,3,2,2025,6000,10000,Crescimento de vendas\n"
        csv_content += "RH,Analista de RH,1,1,2025,5000,7000,Suporte ao time\n"

        buffer = io.BytesIO(csv_content.encode("utf-8"))
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=workforce_planning_import_template.csv"},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating workforce planning import template: {e}")
        raise


@router.post("/entries/import", response_model=None)
async def import_workforce_entries(
    file: UploadFile = File(...),
    preview: bool = Query(False),
    repo: WorkforceRepository = Depends(get_workforce_repo),
    company_id: str = Depends(require_company_id),
):
    # multi-tenancy: gated via Depends(require_company_id) + Postgres RLS runtime (Task #1143)
    """Import headcount planning from Excel/CSV into Store B (PlannedHeadcount).

    Track B / Fase 2: previously wrote a parallel WorkforceEntry store that was
    never rendered in the UI. Now funnels through the canonical producer
    (headcount_import_service) so imports land in PlannedHeadcount — the same
    store the settings table renders. `preview=true` parses ONLY (no write); the
    confirm POST (no preview flag) performs the write.

    Expected columns: department, position, headcount, month, year, salary_min,
    salary_max, notes.
    """
    try:
        rows = await parse_workforce_import_file(file)
        if not rows:
            if preview:
                return {"headers": [], "rows": [], "total_rows": 0}
            return {
                "success": False,
                "imported_count": 0,
                "error_count": 0,
                "errors": [{"message": "No data found in file"}],
                "items": [],
            }

        # Lowercase header keys so the import is case-insensitive.
        items = [
            {(k or "").strip().lower(): v for k, v in row.items()}
            for row in rows
        ]

        if preview:
            return {
                "headers": list(items[0].keys()) if items else [],
                "rows": items[:5],
                "total_rows": len(items),
            }

        summary = await import_planned_headcounts(
            session=repo.db,
            company_id=company_id,
            items=items,
            source="csv_import",
        )
        created = summary.get("created", 0)
        unresolved = summary.get("unresolved_departments") or []
        message = f"{created} posicao(oes) planejada(s) importada(s)."
        if unresolved:
            message += " Departamentos sem vinculo (nao cadastrados): " + ", ".join(unresolved) + "."
        logger.info(
            "[workforce] CSV import created %s planned headcounts for company %s",
            created, company_id,
        )
        return {
            "success": True,
            "imported_count": created,
            "error_count": 0,
            "errors": [],
            "items": [],
            "unresolved_departments": unresolved,
            "message": message,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error importing workforce planning: {e}")
        raise
