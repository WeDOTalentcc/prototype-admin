"""
Job Report Service - PDF and Excel export functionality for job vacancy reports.
"""
import io
import logging
from datetime import datetime
from typing import Any, Literal
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.domains.job_management.repositories.job_report_repository import JobReportRepository
from app.domains.job_management.repositories.job_vacancy_crud_repository import JobVacancyCRUDRepository

from lia_models.candidate import Candidate, VacancyCandidate
from lia_models.job_vacancy import JobVacancy
from lia_models.recruitment_stages import CandidateStageHistory

logger = logging.getLogger(__name__)

ReportFormat = Literal["pdf", "excel"]


class JobReportService:
    """Service for generating job vacancy reports in PDF and Excel formats."""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles for PDF reports."""
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#1a1a2e')
        ))
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.HexColor('#16213e')
        ))
        self.styles.add(ParagraphStyle(
            name='MetricValue',
            parent=self.styles['Normal'],
            fontSize=18,
            textColor=colors.HexColor('#0f3460')
        ))
    
    async def _get_job_vacancy(self, job_id: UUID, company_id: str, db: AsyncSession) -> JobVacancy | None:
        """Get job vacancy with multi-tenancy check."""
        return await JobVacancyCRUDRepository(db).get_by_id_strict_company(
            job_id, company_id
        )
    
    async def _get_funnel_data(self, job_id: UUID, db: AsyncSession) -> dict[str, Any]:
        """Get funnel metrics for a job vacancy."""
        stage_counts_result = await db.execute(
            select(
                VacancyCandidate.stage,
                func.count(VacancyCandidate.id).label("count")
            )
            .where(VacancyCandidate.vacancy_id == job_id)
            .group_by(VacancyCandidate.stage)
        )
        stage_counts = {row.stage: row.count for row in stage_counts_result.all()}
        
        total_result = await db.execute(
            select(func.count(VacancyCandidate.id))
            .where(VacancyCandidate.vacancy_id == job_id)
        )
        total = total_result.scalar() or 0
        
        stage_mapping = {
            'sourcing': ['sourcing'],
            'screening': ['screening', 'triagem', 'initial'],
            'interview': ['interview', 'entrevista'],
            'offer': ['offer', 'proposta'],
            'hired': ['hired', 'contratado'],
            'rejected': ['rejected', 'reprovado', 'recusado']
        }
        
        funnel = {}
        for stage_name, aliases in stage_mapping.items():
            count = sum(stage_counts.get(alias, 0) for alias in aliases)
            funnel[stage_name] = count
        
        return {
            'total': total,
            'stages': funnel,
            'conversion_rates': self._calculate_conversion_rates(funnel, total)
        }
    
    def _calculate_conversion_rates(self, funnel: dict[str, int], total: int) -> dict[str, float]:
        """Calculate conversion rates for each stage."""
        if total == 0:
            return {stage: 0.0 for stage in funnel}
        
        rates = {}
        previous = total
        for stage in ['sourcing', 'screening', 'interview', 'offer', 'hired']:
            count = funnel.get(stage, 0)
            rates[stage] = round((count / previous * 100) if previous > 0 else 0, 1)
            if count > 0:
                previous = count
        return rates
    
    async def _get_source_analysis(self, job_id: UUID, db: AsyncSession) -> list[dict[str, Any]]:
        """Get source distribution for candidates."""
        result = await db.execute(
            select(
                VacancyCandidate.source,
                func.count(VacancyCandidate.id).label("count")
            )
            .where(VacancyCandidate.vacancy_id == job_id)
            .group_by(VacancyCandidate.source)
            .order_by(func.count(VacancyCandidate.id).desc())
        )
        return [{"source": row.source or "Unknown", "count": row.count} for row in result.all()]
    
    async def _get_time_metrics(self, job_id: UUID, db: AsyncSession) -> dict[str, float]:
        """Get average time in each stage."""
        result = await db.execute(
            select(
                CandidateStageHistory.from_stage_name,
                func.avg(CandidateStageHistory.time_in_previous_stage_hours).label("avg_hours")
            )
            .where(CandidateStageHistory.vacancy_id == job_id)
            .group_by(CandidateStageHistory.from_stage_name)
        )
        
        time_metrics = {}
        for row in result.all():
            if row.from_stage_name and row.avg_hours is not None:
                time_metrics[row.from_stage_name] = round(row.avg_hours / 24, 1)
        return time_metrics
    
    async def _get_candidates_list(self, job_id: UUID, db: AsyncSession) -> list[dict[str, Any]]:
        """Get list of candidates for a job vacancy."""
        rows = await JobReportRepository(db).list_candidates_with_profile(job_id)

        candidates = []
        for vc, candidate in rows:
            candidates.append({
                'name': candidate.name,
                'email': candidate.email or '',
                'current_title': candidate.current_title or '',
                'current_company': candidate.current_company or '',
                'stage': vc.stage or 'initial',
                'source': vc.source or 'Unknown',
                'lia_score': vc.lia_score,
                'match_percentage': vc.match_percentage,
                'added_at': vc.created_at.strftime('%Y-%m-%d') if vc.created_at else ''
            })
        return candidates
    
    def _create_pie_chart(self, data: list[dict[str, Any]], width: int = 300, height: int = 200) -> Drawing:
        """Create a pie chart for source distribution."""
        drawing = Drawing(width, height)
        pie = Pie()
        pie.x = 100
        pie.y = 20
        pie.width = 120
        pie.height = 120
        
        if data:
            pie.data = [item['count'] for item in data]
            pie.labels = [item['source'] for item in data]
            
            chart_colors = [
                colors.HexColor('#0f3460'),
                colors.HexColor('#16213e'),
                colors.HexColor('#e94560'),
                colors.HexColor('#1a1a2e'),
                colors.HexColor('#533483'),
                colors.HexColor('#2c3e50'),
            ]
            for i, _ in enumerate(data):
                pie.slices[i].fillColor = chart_colors[i % len(chart_colors)]
        
        drawing.add(pie)
        return drawing
    
    def _create_pdf_table(self, data: list[list[str]], col_widths: list[float] | None = None) -> Table:
        """Create a styled PDF table."""
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16213e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1a1a2e')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        return table
    
    async def generate_funnel_report(
        self,
        job_id: UUID,
        company_id: str,
        format: ReportFormat,
        db: AsyncSession
    ) -> io.BytesIO:
        """Generate funnel metrics report."""
        job = await self._get_job_vacancy(job_id, company_id, db)
        if not job:
            raise ValueError("Job vacancy not found or access denied")
        
        funnel_data = await self._get_funnel_data(job_id, db)
        source_data = await self._get_source_analysis(job_id, db)
        time_metrics = await self._get_time_metrics(job_id, db)
        
        if format == "pdf":
            return self._generate_funnel_pdf(job, funnel_data, source_data, time_metrics)
        else:
            return self._generate_funnel_excel(job, funnel_data, source_data, time_metrics)
    
    def _generate_funnel_pdf(
        self,
        job: JobVacancy,
        funnel_data: dict[str, Any],
        source_data: list[dict[str, Any]],
        time_metrics: dict[str, float]
    ) -> io.BytesIO:
        """Generate PDF funnel report."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
        elements = []
        
        elements.append(Paragraph("📊 Relatório de Funil de Recrutamento", self.styles['ReportTitle']))
        elements.append(Spacer(1, 10))
        
        elements.append(Paragraph(f"<b>Vaga:</b> {job.title}", self.styles['Normal']))
        elements.append(Paragraph(f"<b>Departamento:</b> {job.department or 'N/A'}", self.styles['Normal']))
        elements.append(Paragraph(f"<b>Status:</b> {job.status}", self.styles['Normal']))
        elements.append(Paragraph(f"<b>Gerado em:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", self.styles['Normal']))
        elements.append(Spacer(1, 20))
        
        elements.append(Paragraph("Funil de Candidatos", self.styles['SectionHeader']))
        
        funnel_table_data = [['Etapa', 'Candidatos', 'Taxa de Conversão']]
        stages = funnel_data.get('stages', {})
        rates = funnel_data.get('conversion_rates', {})
        
        stage_labels = {
            'sourcing': 'Sourcing',
            'screening': 'Triagem',
            'interview': 'Entrevista',
            'offer': 'Proposta',
            'hired': 'Contratado',
            'rejected': 'Reprovado'
        }
        
        for stage_key, label in stage_labels.items():
            count = stages.get(stage_key, 0)
            rate = rates.get(stage_key, 0)
            funnel_table_data.append([label, str(count), f"{rate}%"])
        
        funnel_table_data.append(['Total', str(funnel_data.get('total', 0)), '-'])
        elements.append(self._create_pdf_table(funnel_table_data, col_widths=[150, 100, 120]))
        elements.append(Spacer(1, 20))
        
        if source_data:
            elements.append(Paragraph("Distribuição por Fonte", self.styles['SectionHeader']))
            
            source_table_data = [['Fonte', 'Candidatos', 'Percentual']]
            total = sum(s['count'] for s in source_data)
            for source in source_data:
                pct = round(source['count'] / total * 100, 1) if total > 0 else 0
                source_table_data.append([source['source'], str(source['count']), f"{pct}%"])
            
            elements.append(self._create_pdf_table(source_table_data, col_widths=[150, 100, 100]))
            
            if len(source_data) > 1:
                elements.append(Spacer(1, 10))
                elements.append(self._create_pie_chart(source_data))
            
            elements.append(Spacer(1, 20))
        
        if time_metrics:
            elements.append(Paragraph("Tempo Médio por Etapa", self.styles['SectionHeader']))
            
            time_table_data = [['Etapa', 'Dias Médios']]
            for stage, days in time_metrics.items():
                time_table_data.append([stage.capitalize(), f"{days} dias"])
            
            elements.append(self._create_pdf_table(time_table_data, col_widths=[200, 150]))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def _generate_funnel_excel(
        self,
        job: JobVacancy,
        funnel_data: dict[str, Any],
        source_data: list[dict[str, Any]],
        time_metrics: dict[str, float]
    ) -> io.BytesIO:
        """Generate Excel funnel report."""
        wb = Workbook()
        
        ws_summary = wb.active
        ws_summary.title = "Resumo"
        
        header_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws_summary['A1'] = "Relatório de Funil de Recrutamento"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A3'] = "Vaga:"
        ws_summary['B3'] = job.title
        ws_summary['A4'] = "Departamento:"
        ws_summary['B4'] = job.department or 'N/A'
        ws_summary['A5'] = "Status:"
        ws_summary['B5'] = job.status
        ws_summary['A6'] = "Gerado em:"
        ws_summary['B6'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        ws_funnel = wb.create_sheet("Funil")
        
        ws_funnel['A1'] = "Etapa"
        ws_funnel['B1'] = "Candidatos"
        ws_funnel['C1'] = "Taxa de Conversão"
        
        for cell in ws_funnel[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        stages = funnel_data.get('stages', {})
        rates = funnel_data.get('conversion_rates', {})
        stage_labels = {
            'sourcing': 'Sourcing',
            'screening': 'Triagem',
            'interview': 'Entrevista',
            'offer': 'Proposta',
            'hired': 'Contratado',
            'rejected': 'Reprovado'
        }
        
        row = 2
        for stage_key, label in stage_labels.items():
            ws_funnel[f'A{row}'] = label
            ws_funnel[f'B{row}'] = stages.get(stage_key, 0)
            ws_funnel[f'C{row}'] = f"{rates.get(stage_key, 0)}%"
            for col in ['A', 'B', 'C']:
                ws_funnel[f'{col}{row}'].border = border
                ws_funnel[f'{col}{row}'].alignment = Alignment(horizontal='center')
            row += 1
        
        ws_funnel[f'A{row}'] = "Total"
        ws_funnel[f'B{row}'] = funnel_data.get('total', 0)
        ws_funnel[f'C{row}'] = "-"
        for col in ['A', 'B', 'C']:
            ws_funnel[f'{col}{row}'].border = border
            ws_funnel[f'{col}{row}'].font = Font(bold=True)
        
        for col in ['A', 'B', 'C']:
            ws_funnel.column_dimensions[col].width = 20
        
        ws_source = wb.create_sheet("Fontes")
        ws_source['A1'] = "Fonte"
        ws_source['B1'] = "Candidatos"
        ws_source['C1'] = "Percentual"
        
        for cell in ws_source[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        total = sum(s['count'] for s in source_data)
        for i, source in enumerate(source_data, start=2):
            ws_source[f'A{i}'] = source['source']
            ws_source[f'B{i}'] = source['count']
            pct = round(source['count'] / total * 100, 1) if total > 0 else 0
            ws_source[f'C{i}'] = f"{pct}%"
            for col in ['A', 'B', 'C']:
                ws_source[f'{col}{i}'].border = border
        
        for col in ['A', 'B', 'C']:
            ws_source.column_dimensions[col].width = 20
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    async def generate_analytics_report(
        self,
        job_id: UUID,
        company_id: str,
        format: ReportFormat,
        db: AsyncSession
    ) -> io.BytesIO:
        """Generate full analytics report."""
        job = await self._get_job_vacancy(job_id, company_id, db)
        if not job:
            raise ValueError("Job vacancy not found or access denied")
        
        funnel_data = await self._get_funnel_data(job_id, db)
        source_data = await self._get_source_analysis(job_id, db)
        time_metrics = await self._get_time_metrics(job_id, db)
        candidates = await self._get_candidates_list(job_id, db)
        
        if format == "pdf":
            return self._generate_analytics_pdf(job, funnel_data, source_data, time_metrics, candidates)
        else:
            return self._generate_analytics_excel(job, funnel_data, source_data, time_metrics, candidates)
    
    def _generate_analytics_pdf(
        self,
        job: JobVacancy,
        funnel_data: dict[str, Any],
        source_data: list[dict[str, Any]],
        time_metrics: dict[str, float],
        candidates: list[dict[str, Any]]
    ) -> io.BytesIO:
        """Generate PDF analytics report."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
        elements = []
        
        elements.append(Paragraph("📈 Relatório Analítico de Vaga", self.styles['ReportTitle']))
        elements.append(Spacer(1, 10))
        
        elements.append(Paragraph("Informações da Vaga", self.styles['SectionHeader']))
        
        job_info = [
            ['Campo', 'Valor'],
            ['Título', job.title],
            ['Departamento', job.department or 'N/A'],
            ['Localização', job.location or 'N/A'],
            ['Modelo', job.work_model or 'N/A'],
            ['Senioridade', job.seniority_level or 'N/A'],
            ['Status', job.status],
            ['Prioridade', job.priority or 'Média'],
            ['Data Criação', job.created_at.strftime('%d/%m/%Y') if job.created_at else 'N/A'],
        ]
        elements.append(self._create_pdf_table(job_info, col_widths=[150, 300]))
        elements.append(Spacer(1, 20))
        
        elements.append(Paragraph("Métricas do Funil", self.styles['SectionHeader']))
        
        total = funnel_data.get('total', 0)
        stages = funnel_data.get('stages', {})
        hired = stages.get('hired', 0)
        conversion = round(hired / total * 100, 1) if total > 0 else 0
        
        metrics_data = [
            ['Métrica', 'Valor'],
            ['Total de Candidatos', str(total)],
            ['Contratados', str(hired)],
            ['Taxa de Conversão', f"{conversion}%"],
            ['Rejeitados', str(stages.get('rejected', 0))],
        ]
        elements.append(self._create_pdf_table(metrics_data, col_widths=[200, 150]))
        elements.append(Spacer(1, 20))
        
        elements.append(Paragraph("Funil por Etapa", self.styles['SectionHeader']))
        funnel_table = [['Etapa', 'Quantidade', 'Conversão']]
        rates = funnel_data.get('conversion_rates', {})
        for stage_key in ['sourcing', 'screening', 'interview', 'offer', 'hired']:
            label = stage_key.capitalize()
            if stage_key == 'screening':
                label = 'Triagem'
            elif stage_key == 'interview':
                label = 'Entrevista'
            elif stage_key == 'offer':
                label = 'Proposta'
            elif stage_key == 'hired':
                label = 'Contratado'
            
            count = stages.get(stage_key, 0)
            rate = rates.get(stage_key, 0)
            funnel_table.append([label, str(count), f"{rate}%"])
        
        elements.append(self._create_pdf_table(funnel_table, col_widths=[150, 100, 100]))
        elements.append(Spacer(1, 20))
        
        if source_data:
            elements.append(Paragraph("Análise de Fontes", self.styles['SectionHeader']))
            source_table = [['Fonte', 'Candidatos', '%']]
            total_sources = sum(s['count'] for s in source_data)
            for s in source_data[:5]:
                pct = round(s['count'] / total_sources * 100, 1) if total_sources > 0 else 0
                source_table.append([s['source'], str(s['count']), f"{pct}%"])
            elements.append(self._create_pdf_table(source_table, col_widths=[150, 100, 80]))
            
            if len(source_data) > 1:
                elements.append(Spacer(1, 10))
                elements.append(self._create_pie_chart(source_data[:5]))
            elements.append(Spacer(1, 20))
        
        if time_metrics:
            elements.append(Paragraph("Tempo Médio por Etapa", self.styles['SectionHeader']))
            time_table = [['Etapa', 'Dias']]
            for stage, days in time_metrics.items():
                time_table.append([stage.capitalize(), f"{days}"])
            elements.append(self._create_pdf_table(time_table, col_widths=[200, 100]))
            elements.append(Spacer(1, 20))
        
        if candidates:
            elements.append(PageBreak())
            elements.append(Paragraph("Lista de Candidatos (Top 20)", self.styles['SectionHeader']))
            cand_table = [['Nome', 'Cargo Atual', 'Etapa', 'Score']]
            for c in candidates[:20]:
                score = f"{c.get('lia_score', 0):.0f}" if c.get('lia_score') else '-'
                cand_table.append([
                    c.get('name', '')[:30],
                    c.get('current_title', '')[:25],
                    c.get('stage', ''),
                    score
                ])
            elements.append(self._create_pdf_table(cand_table, col_widths=[150, 150, 80, 60]))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def _generate_analytics_excel(
        self,
        job: JobVacancy,
        funnel_data: dict[str, Any],
        source_data: list[dict[str, Any]],
        time_metrics: dict[str, float],
        candidates: list[dict[str, Any]]
    ) -> io.BytesIO:
        """Generate Excel analytics report."""
        wb = Workbook()
        
        header_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws_summary = wb.active
        ws_summary.title = "Resumo"
        
        ws_summary['A1'] = "Relatório Analítico de Vaga"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')
        
        info_data = [
            ('Título', job.title),
            ('Departamento', job.department or 'N/A'),
            ('Localização', job.location or 'N/A'),
            ('Modelo', job.work_model or 'N/A'),
            ('Senioridade', job.seniority_level or 'N/A'),
            ('Status', job.status),
            ('Prioridade', job.priority or 'Média'),
            ('Data Criação', job.created_at.strftime('%d/%m/%Y') if job.created_at else 'N/A'),
        ]
        
        for i, (field, value) in enumerate(info_data, start=3):
            ws_summary[f'A{i}'] = field
            ws_summary[f'B{i}'] = value
            ws_summary[f'A{i}'].font = Font(bold=True)
        
        total = funnel_data.get('total', 0)
        stages = funnel_data.get('stages', {})
        hired = stages.get('hired', 0)
        conversion = round(hired / total * 100, 1) if total > 0 else 0
        
        ws_summary['D3'] = "Métricas Resumidas"
        ws_summary['D3'].font = Font(bold=True, size=12)
        ws_summary['D4'] = "Total Candidatos"
        ws_summary['E4'] = total
        ws_summary['D5'] = "Contratados"
        ws_summary['E5'] = hired
        ws_summary['D6'] = "Taxa Conversão"
        ws_summary['E6'] = f"{conversion}%"
        
        for col in ['A', 'B', 'D', 'E']:
            ws_summary.column_dimensions[col].width = 20
        
        ws_funnel = wb.create_sheet("Funil")
        headers = ['Etapa', 'Candidatos', 'Taxa de Conversão']
        for col, header in enumerate(headers, start=1):
            cell = ws_funnel.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        rates = funnel_data.get('conversion_rates', {})
        stage_labels = [
            ('sourcing', 'Sourcing'),
            ('screening', 'Triagem'),
            ('interview', 'Entrevista'),
            ('offer', 'Proposta'),
            ('hired', 'Contratado'),
            ('rejected', 'Reprovado'),
        ]
        
        for row, (key, label) in enumerate(stage_labels, start=2):
            ws_funnel.cell(row=row, column=1, value=label).border = border
            ws_funnel.cell(row=row, column=2, value=stages.get(key, 0)).border = border
            ws_funnel.cell(row=row, column=3, value=f"{rates.get(key, 0)}%").border = border
        
        for col in range(1, 4):
            ws_funnel.column_dimensions[get_column_letter(col)].width = 18
        
        ws_candidates = wb.create_sheet("Candidatos")
        cand_headers = ['Nome', 'Email', 'Cargo Atual', 'Empresa Atual', 'Etapa', 'Fonte', 'Score LIA', 'Match %', 'Data Adição']
        for col, header in enumerate(cand_headers, start=1):
            cell = ws_candidates.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        for row, c in enumerate(candidates, start=2):
            ws_candidates.cell(row=row, column=1, value=c.get('name', '')).border = border
            ws_candidates.cell(row=row, column=2, value=c.get('email', '')).border = border
            ws_candidates.cell(row=row, column=3, value=c.get('current_title', '')).border = border
            ws_candidates.cell(row=row, column=4, value=c.get('current_company', '')).border = border
            ws_candidates.cell(row=row, column=5, value=c.get('stage', '')).border = border
            ws_candidates.cell(row=row, column=6, value=c.get('source', '')).border = border
            ws_candidates.cell(row=row, column=7, value=c.get('lia_score') or '').border = border
            ws_candidates.cell(row=row, column=8, value=c.get('match_percentage') or '').border = border
            ws_candidates.cell(row=row, column=9, value=c.get('added_at', '')).border = border
        
        col_widths = [25, 30, 25, 25, 15, 15, 12, 12, 15]
        for col, width in enumerate(col_widths, start=1):
            ws_candidates.column_dimensions[get_column_letter(col)].width = width
        
        ws_sources = wb.create_sheet("Fontes")
        source_headers = ['Fonte', 'Candidatos', 'Percentual']
        for col, header in enumerate(source_headers, start=1):
            cell = ws_sources.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        total_sources = sum(s['count'] for s in source_data)
        for row, s in enumerate(source_data, start=2):
            pct = round(s['count'] / total_sources * 100, 1) if total_sources > 0 else 0
            ws_sources.cell(row=row, column=1, value=s['source']).border = border
            ws_sources.cell(row=row, column=2, value=s['count']).border = border
            ws_sources.cell(row=row, column=3, value=f"{pct}%").border = border
        
        for col in range(1, 4):
            ws_sources.column_dimensions[get_column_letter(col)].width = 18
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    async def generate_candidate_list_export(
        self,
        job_id: UUID,
        company_id: str,
        format: ReportFormat,
        db: AsyncSession
    ) -> io.BytesIO:
        """Generate candidate list export."""
        job = await self._get_job_vacancy(job_id, company_id, db)
        if not job:
            raise ValueError("Job vacancy not found or access denied")
        
        candidates = await self._get_candidates_list(job_id, db)
        
        if format == "pdf":
            return self._generate_candidates_pdf(job, candidates)
        else:
            return self._generate_candidates_excel(job, candidates)
    
    def _generate_candidates_pdf(self, job: JobVacancy, candidates: list[dict[str, Any]]) -> io.BytesIO:
        """Generate PDF candidate list."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
        elements = []
        
        elements.append(Paragraph("👥 Lista de Candidatos", self.styles['ReportTitle']))
        elements.append(Paragraph(f"<b>Vaga:</b> {job.title}", self.styles['Normal']))
        elements.append(Paragraph(f"<b>Total:</b> {len(candidates)} candidatos", self.styles['Normal']))
        elements.append(Paragraph(f"<b>Gerado em:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", self.styles['Normal']))
        elements.append(Spacer(1, 20))
        
        if candidates:
            table_data = [['Nome', 'Cargo', 'Empresa', 'Etapa', 'Score']]
            for c in candidates:
                score = f"{c.get('lia_score', 0):.0f}" if c.get('lia_score') else '-'
                table_data.append([
                    c.get('name', '')[:25],
                    c.get('current_title', '')[:20],
                    c.get('current_company', '')[:15],
                    c.get('stage', ''),
                    score
                ])
            
            elements.append(self._create_pdf_table(table_data, col_widths=[120, 110, 100, 80, 50]))
        else:
            elements.append(Paragraph("Nenhum candidato encontrado para esta vaga.", self.styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def _generate_candidates_excel(self, job: JobVacancy, candidates: list[dict[str, Any]]) -> io.BytesIO:
        """Generate Excel candidate list."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidatos"
        
        header_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = f"Lista de Candidatos - {job.title}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:I1')
        
        ws['A2'] = f"Total: {len(candidates)} candidatos | Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A2:I2')
        
        headers = ['Nome', 'Email', 'Cargo Atual', 'Empresa Atual', 'Etapa', 'Fonte', 'Score LIA', 'Match %', 'Data Adição']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        for row, c in enumerate(candidates, start=5):
            ws.cell(row=row, column=1, value=c.get('name', '')).border = border
            ws.cell(row=row, column=2, value=c.get('email', '')).border = border
            ws.cell(row=row, column=3, value=c.get('current_title', '')).border = border
            ws.cell(row=row, column=4, value=c.get('current_company', '')).border = border
            ws.cell(row=row, column=5, value=c.get('stage', '')).border = border
            ws.cell(row=row, column=6, value=c.get('source', '')).border = border
            ws.cell(row=row, column=7, value=c.get('lia_score') or '').border = border
            ws.cell(row=row, column=8, value=c.get('match_percentage') or '').border = border
            ws.cell(row=row, column=9, value=c.get('added_at', '')).border = border
        
        col_widths = [25, 30, 25, 25, 15, 15, 12, 12, 15]
        for col, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


job_report_service = JobReportService()
